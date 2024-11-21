import time     
import logging
import os
import pandas as pd
import numpy as np
import multiprocessing                  #Processing on multiple cores
from functools import partial           #For passing extra arguments to pool.map 
import pymysql
import configparser
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook
from collections import OrderedDict
import json
from packaging.version import LegacyVersion
from pprint import pprint

script_directory = os.path.dirname(os.path.realpath(__file__))

# DB_HOST_IP = '1.21.1.65'
# DB_HOST_IP = '10.110.169.149'
DB_HOST_IP = 'localhost'
DB_USER = 'root'
DB_PASSWD = 'root'
DB_NAME = 'benchtooldb'
DB_PORT = 3306

# The number of cores to be used for multiprocessing
num_processes = 40

# Change the result_type according to result_type_map
# Mapping for Result type field
result_type_map = {0: "single thread", 1: 'single core',
                   2: 'single socket', 3: 'dual socket',
                   4: 'client scaling', 5: '1/8th socket',
                   6: '1/4th socket', 7: '1/2 socket',
                   8: '2 cores', 9: 'perf',
                   10: 'I/O utilization', 11: 'socmon',
                   12: 'OMP_MPI scaling', 20: 'Projection'}

# Returns INPUT_FILTER_CONDITION from 'test_name' and 'input_filters_list'
def get_input_filter_condition(test_name, input_filters_list, wiki_description_file=script_directory + '/config/wiki_description.ini'):
    INPUT_FILTER_CONDITION = ""

    results_metadata_file_path = wiki_description_file
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(results_metadata_file_path)

    # For the input filters
    input_parameters = results_metadata_parser.get(test_name, 'description') \
                                                .replace('\"', '').replace(' ', '').split(',')

    try:
        for index, input_filter in enumerate(input_filters_list):
            if(input_filter != "None"):
                if(input_filter.isnumeric()):
                    INPUT_FILTER_CONDITION += " and SUBSTRING_INDEX(SUBSTRING_INDEX(s.description,','," + str(index+1) +"),',',-1) LIKE \'%" + input_filter  + "%\'"
                else:
                    INPUT_FILTER_CONDITION += " and SUBSTRING_INDEX(SUBSTRING_INDEX(s.description,','," + str(index+1) +"),',',-1) LIKE \'%" + input_filter  + "%\'"
    except Exception as error_message:
        logging.debug("ERRORS::::::: {}".format(error_message))
        pass

    return INPUT_FILTER_CONDITION

# Query the database paralelly for each 'testname'
def parallel_test_report(params, **kwargs):
    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    # Unpack the params here
    test_section, testname, INPUT_FILTER_CONDITION = params

    logging.debug("################################################################################")
    logging.debug("Processing Paralelly for {}".format(testname))
    logging.debug("Input filter condition = {}".format(INPUT_FILTER_CONDITION))

    results_metadata_file_path = script_directory + '/config/wiki_description.ini'
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(results_metadata_file_path)

    description_string = results_metadata_parser[testname]['description'].replace('\"', '')

    SELECT_PARAMS = kwargs['SELECT_PARAMS']
    FINAL_CRITERIA = kwargs['FINAL_CRITERIA']
    kernel_criteria = kwargs['kernel_criteria']
    os_version_criteria = kwargs['os_version_criteria']
    kernel_criteria_op = kwargs['kernel_criteria_op']
    os_version_criteria_op = kwargs['os_version_criteria_op']
    skuid_cpu_map = kwargs['skuid_cpu_map']
    all_skuidnames_criteria = kwargs['all_skuidnames_criteria']
    skuid_criteria_op = kwargs['skuid_criteria_op']
    best_results_condition = kwargs['best_results_condition']

    # An empty dataframe
    results_dataframe = pd.DataFrame()

    if best_results_condition:

        # Get 'higher_is_better' value for the first "field" of testname
        qualifier = results_metadata_parser.get(testname, 'fields').replace('\"', '').split(',')[0]
        higher_is_better = results_metadata_parser.get(testname, 'higher_is_better').replace('\"', '').replace(' ','').split(',')[0]

        # For each skuid_list
        for skuid_list in all_skuidnames_criteria:
            
            # Make SKUID_CRITERIA string
            if skuid_criteria_op == "matches":
                SKUID_CRITERIA = " AND n.skuidname IN " + str(skuid_list).replace('[','(').replace(']',')')
            elif skuid_criteria_op == "does not match":
                SKUID_CRITERIA = " AND n.skuidname NOT IN " + str(skuid_list).replace('[','(').replace(']',')')
            else:
                SKUID_CRITERIA = " AND n.skuidname IN " + str(skuid_list).replace('[','(').replace(']',')')

            # r.number and order by conditions according to higher_is_better
            if higher_is_better == '0':
                R_NUMBER = " MIN(r.number) as number, "
                ORDER_BY_CONDITION = " order by r.number "
            else:
                R_NUMBER = " MAX(r.number) as number, "
                ORDER_BY_CONDITION = " order by r.number DESC "

            if best_results_condition == 'best-results':
                LIMIT_CONDITION = " limit 1 "
            elif best_results_condition == 'top-5':
                LIMIT_CONDITION = " limit 5 "

            QUALIFIER_CONDITION = "  AND disp.qualifier LIKE \'%" + qualifier + '%\''

            GROUP_BY_CONDITION = " group by t.testname, o.originID, " + SELECT_PARAMS + \
                                     " s.description, r.number, disp.unit, disp.qualifier "

            start_time = time.time()

            RESULTS_QUERY = "SELECT t.testname, o.originID, " + SELECT_PARAMS + \
                            "s.description," + R_NUMBER + """ disp.unit, disp.qualifier 
                            FROM result r INNER JOIN subtest s ON s.subtestID=r.subtest_subtestID 
                            INNER JOIN display disp ON disp.displayID=r.display_displayID 
                            INNER JOIN origin o ON o.originID=r.origin_originID 
                            INNER JOIN testdescriptor t ON t.testdescriptorID = o.testdescriptor_testdescriptorID 
                            INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID
                            INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID
                            INNER JOIN node n ON hw.node_nodeID = n.nodeID 
                            INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID
                            INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID
                            WHERE t.testname = \'""" + testname + \
                            "\'" + INPUT_FILTER_CONDITION + FINAL_CRITERIA + SKUID_CRITERIA + " AND r.isvalid = 1 " + \
                            QUALIFIER_CONDITION + GROUP_BY_CONDITION + ORDER_BY_CONDITION + LIMIT_CONDITION + ";"

            logging.debug("\nFINAL QUERY = {}".format(RESULTS_QUERY))

            temp_df = pd.read_sql(RESULTS_QUERY, db)

            query_excecution_time = time.time() - start_time
            logging.debug("Query excecution took {} seconds".format(query_excecution_time))

            # Append the dataframe below the current dataframe
            results_dataframe = results_dataframe.append(temp_df, sort=False)

        # Replace testname with 'test_section' name
        results_dataframe = results_dataframe.reset_index(drop=True)
    else:
        RESULTS_QUERY = "SELECT t.testname, o.originID, " + SELECT_PARAMS + \
                        """s.description, r.number, disp.unit, disp.qualifier 
                        FROM result r INNER JOIN subtest s ON s.subtestID=r.subtest_subtestID 
                        INNER JOIN display disp ON disp.displayID=r.display_displayID 
                        INNER JOIN origin o ON o.originID=r.origin_originID 
                        INNER JOIN testdescriptor t ON t.testdescriptorID = o.testdescriptor_testdescriptorID 
                        INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID
                        INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID
                        INNER JOIN node n ON hw.node_nodeID = n.nodeID 
                        INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID
                        INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID
                        WHERE t.testname = \'""" + testname + \
                        "\'" + INPUT_FILTER_CONDITION + FINAL_CRITERIA + " AND r.isvalid = 1;"

        logging.debug("\nFINAL QUERY = {}".format(RESULTS_QUERY))

        results_dataframe = pd.read_sql(RESULTS_QUERY, db)

    if 'testname' in results_dataframe:
        results_dataframe.insert(0, "Test name", [test_section for item in results_dataframe['testname']])
        del results_dataframe['testname']


    if 'testdate' in results_dataframe.columns:
        results_dataframe.insert(2, 'Test Time', [d.time() for d in results_dataframe['testdate']])
        results_dataframe.insert(2, 'Test Date', [d.date() for d in results_dataframe['testdate']])
        results_dataframe.drop('testdate', axis=1, inplace=True)

    # Filter rows on kernel_criteria
    if kernel_criteria != '':
        if kernel_criteria_op == "greater than":
            logging.debug("greater than selected for ", testname, " Kernel version")
            results_dataframe = results_dataframe[results_dataframe['kernelname'].apply(lambda x: LegacyVersion(x) > LegacyVersion(kernel_criteria))]
        elif kernel_criteria_op == "equals":
            logging.debug("Equals selected  for ", testname, " Kernel version")
            results_dataframe = results_dataframe[results_dataframe['kernelname'].apply(lambda x: LegacyVersion(x) == LegacyVersion(kernel_criteria))]
        else:
            logging.debug("Less than selected!  for ", testname, " Kernel version")
            results_dataframe = results_dataframe[results_dataframe['kernelname'].apply(lambda x: LegacyVersion(x) < LegacyVersion(kernel_criteria))]


    # Filter rows on os_version_criteria
    if os_version_criteria != '':
        if os_version_criteria_op == "greater than":
            logging.debug("greater than selected for ", testname, " OS version")
            results_dataframe = results_dataframe[results_dataframe['osversion'].apply(lambda x: LegacyVersion(x) > LegacyVersion(os_version_criteria))]
        elif kernel_criteria_op == "equals":
            logging.debug("Equals selected for ", testname, " OS version")
            results_dataframe = results_dataframe[results_dataframe['osversion'].apply(lambda x: LegacyVersion(x) == LegacyVersion(os_version_criteria))]
        else:
            logging.debug("Less than selected again!? for ", testname, " OS version")
            results_dataframe = results_dataframe[results_dataframe['osversion'].apply(lambda x: LegacyVersion(x) < LegacyVersion(os_version_criteria))]


    # Corresponding SKUID for skuidname
    # "Unkown SKUID" if skuidname not found in sku_definition.ini (For older benchmarking tests)
    if 'skuidname' in results_dataframe.columns:
        results_dataframe.insert(1, 'SKUID', [skuid_cpu_map.get(skuidname.strip(), "Unkown SKUID") for skuidname in results_dataframe['skuidname']])

    # Convert resultype column into corresponding entry of result_type_map
    if 'resultype' in results_dataframe.columns:
        index = list(results_dataframe.columns).index('resultype')
        results_dataframe.insert(index, 'Result Type', [result_type_map.get(result_type, "Unkown resultype") for result_type in results_dataframe['resultype']])

        # Drop the resultype column as it is no longer needed
        del results_dataframe['resultype']

    if best_results_condition:
        # Do NOT split the description column
        # Since it causes a LOT of columns to be displayed in the final Excel Sheet

        # Rename 'description' to 'Input file Description'
        results_dataframe = results_dataframe.rename(columns = {'description':'Input File Description'})
        pass
    else:
        # SPLIT the description string
        # Add Columns corresponding to the description_string
        description_list = description_string.split(',')
        description_list = list(map(lambda x: x.strip(), description_list))
        for col in reversed(description_list):
            results_dataframe.insert(len(results_dataframe.columns)-4, col, 'default value')

        # Function which splits the description string into various parameters 
        # according to 'description' field of the '.ini' file
        def split_description(index, description):
            try:
                return description.split(',')[index]
            except Exception as error_message:
                logging.debug("Error = {}".format(error_message))
                return np.nan

        # For all the rows in the dataframe, set the description_list values    
        for j in range(len(description_list)):
            results_dataframe[description_list[j]] = results_dataframe['description'].apply(lambda x: split_description(j, x))

        # Drop the 'description' column as we have now split it into various columns according to description_string
        del results_dataframe['description']

    # Add "FACTS Link" column at the end
    index = len(results_dataframe.columns)
    results_dataframe.insert(index, "FACTS Link", ['http://localhost:5000/test-details/' + str(originID) for originID in results_dataframe['originID']])

    return results_dataframe
    
# Get MAX(originID) of given testname paralelly
def parallel_get_max_originID(params, **kwargs):
    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    test_section, testname, INPUT_FILTER_CONDITION = params

    JOIN_SUBTEST = ""
    if INPUT_FILTER_CONDITION:
        JOIN_SUBTEST = " INNER JOIN subtest s on r.subtest_subtestID = s.subtestID "

    ORIGIN_ID_QUERY = """SELECT MAX(o.originID) as originID from origin o 
                        INNER JOIN testdescriptor t ON o.testdescriptor_testdescriptorID = t.testdescriptorID
                        INNER JOIN result r on r.origin_originID = o.originID""" + JOIN_SUBTEST + \
                        " WHERE t.testname = \'" + testname + "\' " + INPUT_FILTER_CONDITION + " AND r.isvalid = 1;"

    latest_originID_df = pd.read_sql(ORIGIN_ID_QUERY, db)

    if not latest_originID_df['originID'][0]:
        return (test_section, 0)
    
    return (test_section, latest_originID_df['originID'][0])

# Get list of test_sections with new originIDs
def get_updated_test_sections(**kwargs):

    base_path = kwargs['base_path']
    all_test_sections = kwargs['all_test_sections']
    all_test_names = kwargs['all_test_names']
    INPUT_FILTER_CONDITION_LIST = kwargs['INPUT_FILTER_CONDITION_LIST']

    # Read the file having test_section -> latest originID read
    reports_last_read_originIDs_file = base_path + 'reports_last_originIDs.txt'

    # If the file doesn't exist or size is ZERO, create a new one with all originIDs = 0
    if not os.path.isfile(reports_last_read_originIDs_file) or os.stat(reports_last_read_originIDs_file).st_size == 0:
        logging.debug("File doesn't exist. Creating it")
        test_originID_map_init = {test_section : 0 for test_section in all_test_sections}
        with open(reports_last_read_originIDs_file, 'w') as f:
            f.write(json.dumps(test_originID_map_init))

    # Read the file and store it in a dictionary
    test_originID_map = {}
    with open(reports_last_read_originIDs_file, 'r') as f:
        test_originID_map = json.loads(f.read())

    logging.debug("Read data from the file. Printing map")

    logging.debug("\n\nExcecuting Queries")
    parallel_start_time = time.time()

    # Get list of test sections for which new originIDs exist (By querying the database) also including input filter condition
    new_originIDs_data = []
    pool = multiprocessing.Pool(num_processes)
    try:
        new_originIDs_data = pool.map(parallel_get_max_originID, zip(all_test_sections, all_test_names, INPUT_FILTER_CONDITION_LIST))
    finally:
        logging.debug("Closing Pool")
        pool.close()
        pool.join()

    logging.debug("Parallelism took {} seconds".format(time.time() - parallel_start_time))

    # test_section -> originID for data just retrieved from the database
    new_originIDs_map = {test_section : int(originID) for test_section, originID in new_originIDs_data}

    # A test_section is 'updated' only if originID from DB > originID read from the file
    updated_test_sections = [test_section for test_section in test_originID_map if new_originIDs_map[test_section] > test_originID_map[test_section]]

    logging.debug("Writing file")   
    logging.debug(new_originIDs_map)
    # Write the updated test_section->originID dictionary to the file
    with open(reports_last_read_originIDs_file, 'w') as f:
        f.write(json.dumps(new_originIDs_map))

    logging.debug("Wrote file returning")
    return updated_test_sections

def get_criteria_params(all_criteria_string, sku_parser, skuid_cpu_map):
    # Metadata 
    parameter_map = {
        "Kernel Version": 'os.kernelname', 
        'OS Version': 'os.osversion', 
        'OS Name': 'os.osdistro', 
        "Firmware Version": 'hw.fwversion' , 
        "ToolChain Name": 'tc.toolchainname', 
        "ToolChain Version" : 'tc.toolchainversion', 
        "Flags": 'tc.flags',
        "SMT" : 'b.smt',
        "Cores": 'b.cores',
        "Corefreq": 'b.corefreq',
        "DDRfreq": 'b.ddrfreq',
        "SKUID": 'n.skuidname',
        "Hostname": 'o.hostname',
        "Scaling" : 's.resultype',
        "Test Date" : 'o.testdate',
        "Notes" : 'o.notes',
    }

    # Required for date criteria
    month_name_number_map = {
        'Jan' : '01',
        'Feb' : '02',
        'Mar' : '03',
        'Apr' : '04',
        'May' : '05',
        'Jun' : '06',
        'Jul' : '07',
        'Aug' : '08',
        'Sep' : '09',
        'Oct' : '10',
        'Nov' : '11',
        'Dec' : '12',
    }


    # Alias for shorter code
    s = all_criteria_string

    best_results_condition = s[8:s.find('-criteria-')]

    advanced_criteria = s[s.find('-criteria-')+len('-criteria-'):s.find('selected-sections-')]

    param_list = json.loads(advanced_criteria)
    selected_sections_list = json.loads(s[s.find('selected-sections-') + len('selected-sections-'):])

    SELECT_PARAMS = " "
    FINAL_CRITERIA = " "
    skuid_criteria_op = ""
    all_skuidnames_criteria = []
    kernel_criteria = ""
    kernel_criteria_op = ""
    os_version_criteria = ""
    os_version_criteria_op = ""

    # Get SELECT_PARAMS and FINAL_CRITERIA
    for d in param_list:
        # Append to SELECT_PARAMS according to 'display' value
        if d['display'] == 'Yes':
            SELECT_PARAMS += parameter_map[d['name']] + ', '

        # Gives 'result type' number from result type string
        # Example if value="2 cores" then key=8
        def get_key_from_value(value):
            return str(list(result_type_map.keys())[list(result_type_map.values()).index(value)])

        # Append FINAL_CRITERIA
        if d['criteria']:
            if d['data_type'] == 'string':
                if d['name'] == 'Scaling':
                    if d['criteria-op'] == 'matches':
                        FINAL_CRITERIA += " AND s.resultype = " + get_key_from_value(d['criteria'].lower())
                    else:
                        FINAL_CRITERIA += " AND s.resultype = " + get_key_from_value(d['criteria'].lower())
                elif d['name'] == 'SKUID' and d['criteria'] != []:
                    for criteria in d['criteria']:
                        # Make list of lists for best_results_condition
                        # Since we have to get best results for each sku
                        if best_results_condition:
                            all_skuidnames_criteria.append(skuid_cpu_map[criteria])
                            skuid_criteria_op = d['criteria-op']
                        else:
                            all_skuidnames_criteria.extend(skuid_cpu_map[criteria])
                            skuid_criteria_op = d['criteria-op']

                    # If not best_results_condition, append the criteria for skuidnames
                    if not best_results_condition:
                        if d['criteria-op'] == 'matches':
                            FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " IN " + str(all_skuidnames_criteria).replace('[','(').replace(']',')')
                        else:
                            FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " NOT IN " + str(all_skuidnames_criteria).replace('[','(').replace(']',')')

                elif d['name'] == 'Kernel Version':
                    kernel_criteria = d['criteria'].strip()
                    kernel_criteria_op = d['criteria-op'].strip()
                elif d['name'] == 'OS Version':
                    os_version_criteria = d['criteria'].strip()
                    os_version_criteria_op = d['criteria'].strip()

                elif d['name'] != 'Kernel Version' and d['name'] != 'OS Version':
                    if d['criteria-op'] == 'matches':
                        FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " LIKE \'%" + d['criteria'].strip() +"%\'"
                    else:
                        FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " NOT LIKE \'%" + d['criteria'].strip() +"%\'"
            elif d['data_type'] == 'numeric':
                if d['criteria-op'] == 'less than':
                    FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " < " + d['criteria'].strip()
                elif d['criteria-op'] == 'equals':
                    FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " = " + d['criteria'].strip()
                elif d['criteria-op'] == 'greater than':
                    FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " > " + d['criteria'].strip()
            elif d['data_type'] == 'date':
                if d['criteria-op'] == 'before':
                    FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " < \'" + d['criteria'].strip() + '-' + month_name_number_map[d['criteria2']] + '-' + '01' + "\'"
                elif d['criteria-op'] == 'during':
                    FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " LIKE \'%" + d['criteria'].strip() + '-' + month_name_number_map[d['criteria2']] + "-" "%\'"
                elif d['criteria-op'] == 'since':
                    FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " > \'" + d['criteria'].strip() + '-' + month_name_number_map[d['criteria2']] + '-' + '01' + "\'"

    # Handle the condition where best_results_condition exists AND 'all_skuidnames_criteria' is empty []
    if best_results_condition and all_skuidnames_criteria == []:
        for criteria in sku_parser.sections():
            all_skuidnames_criteria.append(skuid_cpu_map[criteria])

    return {
        'SELECT_PARAMS' : SELECT_PARAMS, 
        'FINAL_CRITERIA' : FINAL_CRITERIA, 
        'kernel_criteria' : kernel_criteria, 
        'kernel_criteria_op' : kernel_criteria_op, 
        'os_version_criteria' : os_version_criteria, 
        'os_version_criteria_op' : os_version_criteria_op, 
        'best_results_condition' : best_results_condition, 
        'skuid_criteria_op' : skuid_criteria_op,
        'all_skuidnames_criteria' : all_skuidnames_criteria,
        'selected_sections_list' : selected_sections_list
    }

# Update all cached reports (.xlsx) in 'cached_results' directory 
def update_cached_reports():
    results_metadata_file_path = script_directory + '/config/best_of_all_graph.ini'
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(results_metadata_file_path)

    sku_file_path = script_directory + '/config/sku_definition.ini'
    sku_parser = configparser.ConfigParser()
    sku_parser.read(sku_file_path)

    #  Map from CPU Section->list(skuidname) in sku_definition.ini 
    skuid_cpu_map = OrderedDict({section: sku_parser.get(section, 'SKUID').replace('\"', '').split(',') for section in sku_parser.sections()})

    # Also a map from skuidname->Cpu Section. We need both in diff scenarios
    for section in sku_parser.sections():
        skuids = sku_parser.get(section, 'SKUID').replace('\"','').split(',')
        for skuid in skuids:
            skuid_cpu_map[skuid] = section

    # Map from filter label -> testname_list Eg. ncar -> ['clubb', 'mg2', 'waccm', 'wrf']
    label_testname_map = {}
    for section in results_metadata_parser.sections():
        labels = results_metadata_parser.get(section, 'label').replace('\"','').replace(' ', '').lower().split(',')

        # Remove empty strings
        labels = list(filter(None, labels))
        for label in labels:
            # If the list doesn't exist, create one
            if label not in label_testname_map:
                label_testname_map[label] = []
            label_testname_map[label].append(section)

    # Get all test_sections from 'best_of_all_graph.ini'
    all_test_sections = results_metadata_parser.sections()

    # Sort the selected_tests_list alphabetically
    all_test_sections.sort(key=str.lower)
    # Get the INPUT_FILTER_CONDITION for each selected test
    input_filters_list_list = [results_metadata_parser.get(test_section, 'default_input') \
                                .replace('\"', '').split(',') for test_section in all_test_sections]
    INPUT_FILTER_CONDITION_LIST = [get_input_filter_condition(test_section, input_filters_list, \
                                    wiki_description_file= script_directory + "/config/best_of_all_graph.ini") \
                                    for test_section, input_filters_list in zip(all_test_sections, input_filters_list_list)]

    all_test_names = [results_metadata_parser.get(section, 'testname').strip() for section in all_test_sections] 

    # The directory where cached excel files will be stored
    base_path = script_directory + '/cached_results/'
    # If directory doesn't exist, create it
    if not os.path.exists(base_path):
        os.mkdir(base_path)         
        
    updated_test_sections = get_updated_test_sections(base_path=base_path, all_test_sections=all_test_sections, \
                            all_test_names=all_test_names, INPUT_FILTER_CONDITION_LIST=INPUT_FILTER_CONDITION_LIST)

    # Read the reports_cache_map and regenerate the queries
    # Check if the excel file path for the current request exists in the map_file
    reports_cache_map_file = base_path + 'reports_cache_map.txt'

    # If reports_cache_map.txt doesn't exist, create it again
    if not os.path.isfile(reports_cache_map_file):
        os.system("touch " + reports_cache_map_file)

    # The actual dictionary. Fill it by using json.loads()
    reports_cache_map = {}
    with open(reports_cache_map_file, 'r') as f:
        try:
            reports_cache_map = json.loads(f.read())
        except:
            # In case the file is empty, ignore
            pass

    # For each cache entry in reports_cache_map
    for all_criteria_string in reports_cache_map:
        absolute_file_path = reports_cache_map[all_criteria_string]

        # If the file exists, then update it for updated_test_sections
        if os.path.exists(absolute_file_path):

            # Get all params for recreating the query from all_criteria_string
            criteria_params = get_criteria_params(all_criteria_string, sku_parser, skuid_cpu_map)

            SELECT_PARAMS = criteria_params['SELECT_PARAMS']
            FINAL_CRITERIA = criteria_params['FINAL_CRITERIA']
            kernel_criteria = criteria_params['kernel_criteria']
            kernel_criteria_op = criteria_params['kernel_criteria_op']
            os_version_criteria = criteria_params['os_version_criteria']
            os_version_criteria_op = criteria_params['os_version_criteria_op']
            best_results_condition = criteria_params['best_results_condition']
            skuid_criteria_op = criteria_params['skuid_criteria_op']
            all_skuidnames_criteria = criteria_params['all_skuidnames_criteria']
            selected_sections_list = criteria_params['selected_sections_list']

            # Filter selected_sections by according to updated_test_sections
            if not best_results_condition:
                selected_sections_list = [x for x in selected_sections_list if x in updated_test_sections]
            selected_tests_list = [results_metadata_parser.get(section, 'testname').strip() for section in selected_sections_list]

            input_filters_list_list = [results_metadata_parser.get(test_section, 'default_input') \
                                        .replace('\"', '').split(',') for test_section in selected_sections_list]
            INPUT_FILTER_CONDITION_LIST = [get_input_filter_condition(test_section, input_filters_list, \
                                            wiki_description_file=script_directory +"/config/best_of_all_graph.ini") \
                                            for test_section, input_filters_list in zip(updated_test_sections, input_filters_list_list)]

            logging.debug(SELECT_PARAMS)
            logging.debug(FINAL_CRITERIA)
            logging.debug(selected_sections_list)

            # If selected_sections_list is not empty, i.e. if atleast one test is updated
            if selected_sections_list:
                parallel_start_time = time.time()

                # Parallel excecution 
                results_dataframe_list = []
                pool = multiprocessing.Pool(num_processes)
                try:
                    results_dataframe_list = pool.map(partial(parallel_test_report, SELECT_PARAMS=SELECT_PARAMS, FINAL_CRITERIA=FINAL_CRITERIA, \
                                            kernel_criteria=kernel_criteria, kernel_criteria_op=kernel_criteria_op, \
                                            os_version_criteria=os_version_criteria, os_version_criteria_op=os_version_criteria_op, \
                                            skuid_cpu_map=skuid_cpu_map, best_results_condition=best_results_condition, skuid_criteria_op=skuid_criteria_op, \
                                            all_skuidnames_criteria=all_skuidnames_criteria), zip(selected_sections_list, selected_tests_list, INPUT_FILTER_CONDITION_LIST))
                finally:
                    logging.debug("Closing Pool")
                    pool.close()
                    pool.join()

                logging.debug("Parallelism took {} seconds".format(time.time() - parallel_start_time))

                start_time2 = time.time()

                logging.debug("Writing excel sheets")
                # Write all results in a single excel file
                if best_results_condition:
                    logging.debug("Best results. Writing excel sheets")
                    # Write all results in a single excel file

                    final_results_dataframe = pd.DataFrame()

                    # Append all the dataframes from the list to final_results_dataframe
                    i = 0
                    for results_dataframe in results_dataframe_list:
                        i += 1
                        final_results_dataframe = final_results_dataframe.append(results_dataframe, sort=False)

                        # Append an empty row to the final_results_dataframe
                        final_results_dataframe = final_results_dataframe.append(pd.Series(), ignore_index=True, sort=False)

                    # Reset Index
                    final_results_dataframe = final_results_dataframe.reset_index(drop=True)

                    logging.debug("final results dataframe = {}".format(i))

                    logging.debug("Final results dataframe = ")
                    logging.debug(final_results_dataframe)

                    # Write the entire dataframe in a single sheet in append mode
                    with pd.ExcelWriter(absolute_file_path, engine='openpyxl') as writer:
                        final_results_dataframe.to_excel(writer, sheet_name="Best results")

                else:
                    wb = load_workbook(absolute_file_path)
                    for results_dataframe, section in zip(results_dataframe_list, selected_sections_list):
                        if section in wb.sheetnames:
                            # Get index of that sheet
                            index = wb.sheetnames.index(section)

                            # Delete the existing sheet
                            del wb[section]

                            # Insert the new sheet at 'index' position
                            ws = wb.create_sheet(section, index)

                            for r in dataframe_to_rows(results_dataframe, index=True, header=True):
                                ws.append(r)

                    wb.save(absolute_file_path)
                logging.debug("Writing all excel files took {} seconds".format(time.time() - start_time2))
        else:
            # Delete the entry from the cache map
            del reports_cache_map[all_criteria_string]

    # Write the updated reports_cache_map_file back to the file
    with open(reports_cache_map_file, 'w') as f:
        f.write(json.dumps(reports_cache_map, indent=4))


if __name__ == "__main__":
    print("Running script")

    start_time = time.time()
    update_cached_reports()

    print("Done running the script in {} seconds".format(time.time() - start_time))
