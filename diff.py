from pprint import pprint
import time
from datetime import datetime
import logging
import os, shutil
import multiprocessing                  #Processing on multiple cores
from functools import partial           #For passing extra arguments to pool.map 
import pandas as pd
import numpy as np
import pymysql
import configparser, itertools
from flask import Flask, render_template, request, redirect, send_file, url_for, session, send_from_directory
from collections import OrderedDict
import csv
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from packaging.version import LegacyVersion
import json

# DB_HOST_IP = '1.21.1.65'
# DB_HOST_IP = '10.110.169.149'
DB_HOST_IP = 'localhost'
DB_USER = 'root'
DB_PASSWD = 'root'
DB_NAME = 'benchtooldb'
DB_PORT = 3306

# The number of cores to be used for multiprocessing
num_processes = 40

# Change the result_type according to result_type_map
# Mapping for Result type field
result_type_map = {0: "single thread", 1: 'single core',
                   2: 'single socket', 3: 'dual socket',
                   4: 'client scaling', 5: '1/8th socket',
                   6: '1/4th socket', 7: '1/2 socket',
                   8: '2 cores', 9: 'perf',
                   10: 'I/O utilization', 11: 'socmon',
                   12: 'OMP_MPI scaling', 20: 'Projection'}

month_name_map = {1:'Jan', 2:'Feb', 3:'Mar', 4:'Apr', 5:'May', 6:'Jun', 7:'Jul', 8:'Aug', 9:'Sep', 10: 'Oct', 11:'Nov', 12:'Dec'}

# Uncomment this line for toggling debugging messages on the console
logging.basicConfig(level=logging.DEBUG)

app = Flask(__name__)

logging.debug("Flask server restarted")

# Just a random secret key. Created by md5 hashing the string 'secretactividad'
app.secret_key = "05ec4a13767ac57407c4000e55bdc32c"
pd.set_option('display.max_rows', 500)

# RETURNS the Table name for the given 'index' from the dictionary of lists
# example : from 'origin_param_list' -> return 'origin'
@app.context_processor
def table_name():
    def _table_name(list_of_keys, index):
        tablename = list(list_of_keys)[index]
        if tablename == "ram_details_param_list":
            return "RAM_details"
        return tablename[0: tablename.find("_param_list")].capitalize()

    return dict(table_name=_table_name)


# RESERVED FOR LATER (Not important as of now. DO it if time permits)
# @app.template_filter('readable_timestamp')
# def readable_timestamp(timestamp):
#     #type of timestamp is <class 'pandas._libs.tslibs.timestamps.Timestamp'>
#     dt = timestamp.to_pydatetime();
#     return dt.strftime("%d %B, %Y %I:%M:%S %p")

# Takes the originID, returns the testname
def get_test_name(originID):
    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    TEST_NAME_QUERY = """SELECT t.testname FROM testdescriptor as t 
                        INNER JOIN origin o on o.testdescriptor_testdescriptorID=t.testdescriptorID 
                        WHERE o.originID=""" + originID + ";"

    test_name_dataframe = pd.read_sql(TEST_NAME_QUERY, db)
    test_name = test_name_dataframe['testname'][0]

    # close the database connection
    try:
        db.close()
    except:
        pass

    return test_name

#Template filter for 'unique_list'
@app.template_filter('unique_list')
def unique_list_filter(input_list):
    return unique_list(input_list)

# Takes input as a list containg duplicate elements. 
# Returns a sorted list having unique elements
def unique_list(input_list, reverse=False):
    def str_is_int(s):
        try:
            int(s)
            return True
        except:
            return False

    def str_is_float(s):
        try:
            # This fails if string isn't float
            float(s) == int(s)
            return True
        except:
            return False

    # OrderedDict creates unique keys. It also preserves the order of insertion
    lst = list(OrderedDict.fromkeys(input_list))

    logging.debug(" = {}".format(lst))

    if all(str_is_int(x) for x in lst):
        logging.debug("WAS INSTANCE OF INT MAN")
        lst = [int(x) for x in lst]
    elif all(str_is_float(x) for x in lst):
        logging.debug("WAS INSTANCE OF FLOAT MAN")
        lst = [float(x) for x in lst]
    else:
        logging.debug("WAS INSTANCE OF NONE MAN")

    # Return all values as STR
    if reverse:
        return list(reversed(list(map(lambda x: str(x), sorted(lst)))))
    else:
        return list(map(lambda x: str(x), sorted(lst)))

@app.template_filter('no_of_rows')
def no_of_rows(dictionary):
    if dictionary != {}:
        # this is a dictionary of lists
        # return length of the first list in the dictionary
        # fastest way
        return len(dictionary[next(iter(dictionary))])
    else:
        return 0

def read_all_parameter_lists(parameter_lists, test_name):

    # read metadata from metadata.ini file
    env_metadata_file_path = './config/metadata.ini'
    env_metadata_parser = configparser.ConfigParser()
    env_metadata_parser.read(env_metadata_file_path)

    # Read metadata for results in wiki_description.ini file
    results_metadata_file_path = './config/wiki_description.ini'
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(results_metadata_file_path)

    # Fill all parameter Lists in the dictionary
    for param_list_name in parameter_lists:
        if param_list_name == 'results_param_list':
            parameter_lists[param_list_name] = results_metadata_parser.get(test_name, 'description') \
                                                .replace('\"', '').replace(' ', '').split(',')
            parameter_lists[param_list_name].extend(['number', 'resultype', 'unit', 'qualifier'])

        elif param_list_name == 'qualifier':
            parameter_lists[param_list_name] = results_metadata_parser.get(test_name, 'fields') \
                                                .replace('\"','').lower().split(',')

        elif param_list_name == 'min_or_max':
            parameter_lists[param_list_name] = results_metadata_parser.get(test_name, 'higher_is_better') \
                                                .replace('\"','').split(',')

        else:
            # extracts 'example' from 'example_param_list'
            env_param_name = param_list_name[0:param_list_name.find("_param_list")]

            parameter_lists[param_list_name] = env_metadata_parser.get(env_param_name, 'db_variables') \
                                                .replace(' ', '').split(',')
    return parameter_lists

def read_all_csv_files(compare_lists, parameter_lists, originID_compare_list):
    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    JENKINS_QUERY = """SELECT J.jobname, J.runID FROM origin O
                        INNER JOIN jenkins J ON O.jenkins_jenkinsID=J.jenkinsID 
                        AND O.originID in """ + str(originID_compare_list).replace('[', '(').replace(']', ')') + ";"
    jenkins_details = pd.read_sql(JENKINS_QUERY, db)

    jobname_list = jenkins_details['jobname'].to_list()
    runID_list = jenkins_details['runID'].to_list()

    # The path on which file is to be read
    file_path = "/mnt/nas/dbresults/"

    # Fill all lists in the dictionary
    for i in range(len(originID_compare_list)):
        for j in range(len(compare_lists)):
            param_values_dictionary = dict()
            list_of_keys = list(compare_lists.keys())
            table_name = list_of_keys[j][0:list_of_keys[j].find("_list")]

            # Check if the file exists
            if (
            os.path.exists(file_path + str(jobname_list[i]) + '/' + str(runID_list[i]) + '/' + table_name + '.csv')):
                pass
            else:
                table_name = list_of_keys[j][0:list_of_keys[j].find("_details_list")]
                if(os.path.exists(file_path + str(jobname_list[i]) + '/' + str(runID_list[i]) + '/' + table_name + '.csv')):
                    pass
                else:
                    continue
            with open(file_path + str(jobname_list[i]) + '/' + str(runID_list[i]) + '/' + table_name + '.csv',
                      newline='') as csvfile:
                reader = csv.reader(csvfile, delimiter=',')
                for row in reader:
                    try:
                        for k in range(0, len(parameter_lists[table_name + '_param_list'])):
                            param_values_dictionary[parameter_lists[table_name + '_param_list'][k]] = row[k]
                        break
                    except IndexError:
                        #If Index error occurs, pass. Let the value remain empty
                        pass
                    except:
                        for k in range(0, len(parameter_lists[table_name + '_details_param_list'])):
                            param_values_dictionary[parameter_lists[table_name + '_details_param_list'][k]] = row[k]
                        break
            if (table_name + "_list" in compare_lists):
                compare_lists[table_name + "_list"].append(param_values_dictionary)
            else:
                compare_lists[table_name + "_details_list"].append(param_values_dictionary)
 
     # close the database connection
    try:
        db.close()
    except:
        pass

    return compare_lists

# Returns INPUT_FILTER_CONDITION from 'test_name' and 'input_filters_list'
def get_input_filter_condition(test_name, input_filters_list, wiki_description_file='./config/wiki_description.ini'):
    INPUT_FILTER_CONDITION = ""

    results_metadata_file_path = wiki_description_file
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(results_metadata_file_path)

    # For the input filters
    input_parameters = results_metadata_parser.get(test_name, 'description') \
                                                .replace('\"', '').replace(' ', '').split(',')

    try:
        for index, input_filter in enumerate(input_filters_list):
            if(input_filter != "None"):
                if(input_filter.isnumeric()):
                    INPUT_FILTER_CONDITION += " and SUBSTRING_INDEX(SUBSTRING_INDEX(s.description,','," + str(index+1) +"),',',-1) = \'" + input_filter  + "\'"
                else:
                    INPUT_FILTER_CONDITION += " and SUBSTRING_INDEX(SUBSTRING_INDEX(s.description,','," + str(index+1) +"),',',-1) LIKE \'%" + input_filter  + "%\'"
    except Exception as error_message:
        logging.debug("ERRORS::::::: {}".format(error_message))
        pass

    return INPUT_FILTER_CONDITION

# Request route for favicon
@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'),
                               'images/favicon.ico', mimetype='image/vnd.microsoft.icon')

# Get all-tests data
def get_all_tests_data(wiki_description_file='./config/wiki_description.ini'):
    parser = configparser.ConfigParser()
    parser.read(wiki_description_file)

    # Reference for best_of_all_graph
    sku_file_path = './config/sku_definition.ini'
    sku_parser = configparser.ConfigParser()
    sku_parser.read(sku_file_path)

    reference_list = sku_parser.sections();
    logging.debug("SECTIONS")
    logging.debug("{}".format(reference_list))


    filter_labels_dict = {}
    filter_labels_list = []
    hpc_benchmarks_list = []
    cloud_benchmarks_list = []
    
    for section in parser.sections():
        filter_labels_list.extend(
            [label for label in parser.get(section, 'label').replace('\"', '').replace(' ', '').lower().split(',')
            if label != '' and label not in filter_labels_list])
        filter_labels_dict[section] = ','.join(
            [label for label in parser.get(section, 'label').replace('\"', '').replace(' ', '').lower().split(',')])
        
        type_of_benchmark = parser.get(section, 'model').strip()
        
        if type_of_benchmark == '\"hpc\"':
            hpc_benchmarks_list.append(section)
        else:
            cloud_benchmarks_list.append(section)


    hpc_benchmarks_list = sorted(hpc_benchmarks_list, key=str.lower)
    cloud_benchmarks_list = sorted(cloud_benchmarks_list, key=str.lower)
    filter_labels_list = sorted(filter_labels_list, key=str.lower)

    # If wiki_description_file = best_of_all_graph.ini, then benchmark list is actually sections list
    # Give its value to sections_list and update benchmarks_list

    hpc_sections_list = []
    cloud_sections_list = []
    if wiki_description_file == './config/best_of_all_graph.ini':
        hpc_sections_list, cloud_sections_list = hpc_benchmarks_list, cloud_benchmarks_list
        hpc_benchmarks_list = [parser.get(section, 'testname').strip() for section in hpc_sections_list]
        cloud_benchmarks_list = [parser.get(section, 'testname').strip() for section in cloud_sections_list]

        # Unique entries only 
        hpc_benchmarks_list = sorted(list(set(hpc_benchmarks_list)), key=str.lower)
        cloud_benchmarks_list = sorted(list(set(cloud_benchmarks_list)), key=str.lower)

    context = {
        'hpc_benchmarks_list': hpc_benchmarks_list,
        'cloud_benchmarks_list': cloud_benchmarks_list,
        'filter_labels_list': filter_labels_list,
        'filter_labels_dict': filter_labels_dict,
        'reference_list': reference_list,
        'hpc_sections_list' : hpc_sections_list,
        'cloud_sections_list' : cloud_sections_list,
    }

    return context

# Error 404 custom page not found
@app.errorhandler(404)
def page_not_found(e):
    # For 'Go To Benchmark' Dropdown
    all_tests_data = get_all_tests_data()
    
    # note that we set the 404 status explicitly
    return render_template('404.html', all_tests_data=all_tests_data), 404

# ALL TESTS PAGE
@app.route('/')
def home_page():
    context = get_all_tests_data()

    # For result type filter
    result_type_list = [3, 2, 7, 6, 5, 8, 1, 0]
    # Convert to result_type string according to result_type_map
    result_type_list = [result_type_map[x] for x in result_type_list]

    context['result_type_list'] = result_type_list

    return render_template('all-tests.html', context=context)

@app.route('/about')
def about_page():
    # For 'Go To Benchmark' Dropdown
    all_tests_data = get_all_tests_data()

    return render_template('about.html', context = {}, all_tests_data=all_tests_data)

# Get data for All runs of the test 'testname' from database
def get_all_runs_data(testname, secret=False):
    # Read metadata for results in wiki_description.ini file
    results_metadata_file_path = './config/wiki_description.ini'
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(results_metadata_file_path)

    qualifier_list = results_metadata_parser.get(testname, 'fields').replace('\"', '').split(',')
    min_or_max_list = results_metadata_parser.get(testname, 'higher_is_better') \
                    .replace('\"', '').replace(' ', '').split(',')

    logging.debug("########PRINTING QUALIFIER LIST AND MIN OR MAX LIST#########")
    logging.debug("{}".format(qualifier_list))
    logging.debug("{}".format(min_or_max_list))

    if secret == True:
        RESULTS_VALIDITY_CONDITION = " "
    else:
        RESULTS_VALIDITY_CONDITION = " AND r.isvalid = 1 "

    if min_or_max_list[0] == '0':
        ALL_RUNS_QUERY = "SELECT DISTINCT o.originID, o.testdate, o.hostname, MIN(r.number) as \'Best" +\
                        qualifier_list[0].replace(" ",'') + """\', o.notes, r.isvalid from result r INNER JOIN display disp 
                        ON  r.display_displayID = disp.displayID
                        INNER JOIN origin o ON o.originID = r.origin_originID 
                        INNER JOIN testdescriptor t ON t.testdescriptorID = o.testdescriptor_testdescriptorID 
                        where t.testname = \'""" + testname + """\' 
                        AND disp.qualifier LIKE \'%""" + qualifier_list[0] + "%\'" + \
                        RESULTS_VALIDITY_CONDITION + """ GROUP BY o.originID, o.testdate, o.hostname, o.notes, r.isvalid
                        ORDER BY o.originID DESC"""
    else:
        ALL_RUNS_QUERY = "SELECT DISTINCT o.originID, o.testdate, o.hostname, MAX(r.number) as \'Best" +\
                        qualifier_list[0].replace(" ",'') + """\', o.notes, r.isvalid from result r INNER JOIN display disp 
                        ON  r.display_displayID = disp.displayID 
                        INNER JOIN origin o ON o.originID = r.origin_originID 
                        INNER JOIN testdescriptor t ON t.testdescriptorID = o.testdescriptor_testdescriptorID 
                        where t.testname = \'""" + testname + """\' 
                        AND disp.qualifier LIKE \'%""" + qualifier_list[0] + "%\'" + \
                        RESULTS_VALIDITY_CONDITION + """ GROUP BY o.originID, o.testdate, o.hostname, o.notes, r.isvalid
                        ORDER BY o.originID DESC"""
    
    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    dataframe = pd.read_sql(ALL_RUNS_QUERY, db)
    rows, columns = dataframe.shape  # returns a tuple (rows,columns)

    # For secret page, return only table data. The secret function will redirect to secret page
    if secret == True:
        secret_context = {
            'testname': testname,
            'data': dataframe.to_dict(orient='list'),
            'no_of_rows': rows,
            'no_of_columns': columns,
        }

        # close the database connection
        try:
            db.close()
        except:
            pass
        return secret_context
    # Else render all-runs.html
    else:
        del dataframe['isvalid']
        rows, columns = dataframe.shape  # returns a tuple (rows,columns)

        # Dropdown for input file
        input_parameters = results_metadata_parser.get(testname, 'description') \
                                                    .replace('\"', '').replace(' ', '').split(',')

        INPUT_FILE_QUERY = """SELECT DISTINCT s.description, r.isvalid FROM origin o INNER JOIN testdescriptor t
                                ON t.testdescriptorID=o.testdescriptor_testdescriptorID  INNER JOIN result r
                                ON o.originID = r.origin_originID INNER JOIN subtest s
                                ON  r.subtest_subtestID = s.subtestID WHERE t.testname = \'""" + testname + "\';" #RRG
        logging.debug(INPUT_FILE_QUERY)
        try:
            input_details_df = pd.read_sql(INPUT_FILE_QUERY, db)
        except:
            pass
        finally:
            db.close()

        # Filter results which are valid
        input_details_df = input_details_df[input_details_df['isvalid'] == 1].reset_index(drop=True)
        del input_details_df['isvalid']
        
        logging.debug("{}".format(input_parameters))
        logging.debug("{}".format(input_details_df))

        # Function which splits the description string into various parameters 
        # according to 'description' field of the '.ini' file
        def split_description(index, description):
            try:
                return description.split(',')[index]
            except Exception as error_message:
                logging.debug("{}".format(error_message))
                return np.nan

        # Split the 'description' column into multiple columns
        for index, param in enumerate(input_parameters):
            input_details_df[param] = input_details_df['description'].apply(lambda x: split_description(index, x))

        # Delete the 'description' column
        del input_details_df['description']
        
        # Drop all the rows which have NaN as an element
        input_details_df.dropna(inplace=True)
        
        # Get default_inputs from wiki_description.ini
        default_input_filters_list = results_metadata_parser.get(testname, 'default_input') \
                                                .replace('\"', '').split(',')

        # Read from test_summary.ini file
        test_summary_file_path = "./config/test_summary.ini"
        test_summary_parser = configparser.ConfigParser()
        test_summary_parser.read(test_summary_file_path)

        test_summary = {}
        test_summary['summary'] = test_summary_parser.get(testname, 'summary').replace('\"','')
        test_summary['source_code_link'] = test_summary_parser.get(testname, 'source_code_link')
        test_summary['type_of_workload'] = test_summary_parser.get(testname, 'type_of_workload')
        test_summary['default_input'] = test_summary_parser.get(testname, 'default_input')
        test_summary['latest_version'] = test_summary_parser.get(testname, 'latest_version')

        # For result type filter
        result_type_list = [3, 2, 7, 6, 5, 8, 1, 0]
        # Convert to result_type string according to result_type_map
        result_type_list = [result_type_map[x] for x in result_type_list]

        context = {
            'testname': testname,
            'data': dataframe.to_dict(orient='list'),
            'no_of_rows': rows,
            'no_of_columns': columns,
            'qualifier_list': qualifier_list,
            'input_details': input_details_df.to_dict(orient='list'),
            'default_input_filters': default_input_filters_list,
            'test_summary' : test_summary,
            'result_type_list' : result_type_list,
        }

        # close the database connection
        try:
            db.close()
        except:
            pass

        return context

# Show all runs of a test 'testname'
@app.route('/allruns/<testname>', methods=['GET'])
def all_runs_page(testname):
    
    # Reference for best_of_all_graph
    sku_file_path = './config/sku_definition.ini'
    sku_parser = configparser.ConfigParser()
    sku_parser.read(sku_file_path)


    # Reference dropdown for timeline graphs
    all_skus_list = sku_parser.sections()

    try:
        context = get_all_runs_data(testname)
        context['all_skus_list'] = all_skus_list
        error = None
    except Exception as error_message:
        context = None
        error = error_message

    # For 'Go To Benchmark' Dropdown
    all_tests_data = get_all_tests_data()

    return render_template('all-runs.html', error=error, context=context, all_tests_data=all_tests_data)

# Page for marking a test 'originID' invalid
@app.route('/allruns/secret/<testname>', methods=['GET', 'POST'])
def all_runs_secret_page(testname):
    if request.method == 'GET':
        # For 'Go To Benchmark' Dropdown
        all_tests_data = get_all_tests_data()

        return render_template('secret-all-runs.html', testname={'name':testname}, context={}, all_tests_data=all_tests_data)
    else:
        success = {}
        error = {}
        keyerror = {}
        logging.debug("#######POSTED###########")
        logging.debug("{}".format(request.args))
        
        # Get doesn't throw error. 
        # If key is not present it sets to default ('None' most of the times)
        success = session.get('success')
        error = session.get('error')
        keyerror = session.get('keyerror')
    
        logging.debug('success = {}'.format(success))
        logging.debug('error = {}'.format(error))
        logging.debug('keyerror = {}'.format(keyerror))

        logging.debug("#######SESSION BEFORE ####")
        logging.debug(' = {}'.format(session))
        logging.debug("########SESSION AFTER $$$$")
        # Clear the contents of the session (cookies)
        session.clear()
        logging.debug(' = {}'.format(session))

        context = context = get_all_runs_data(testname, secret=True)

        # For 'Go To Benchmark' Dropdown
        all_tests_data = get_all_tests_data()

        return render_template('secret-all-runs.html', success=success, error=error, keyerror=keyerror, context=context, all_tests_data=all_tests_data)

@app.route('/mark-origin-id-invalid', methods=['POST'])
def mark_originID_invalid():
    logging.debug("\n\n\n#REQUEST#########")
    logging.debug(" = {}".format(request.form))

    data = json.loads(request.form.get('data'))
    logging.debug(" = {}".format(data))

    originIDs = data.get('originIDs')
    logging.debug("Printing selected originIDs = '{}' {}".format(originIDs, type(originIDs)))

    testname = data.get('testname')
    valid = data.get('valid')
    secret_key = data.get('secretKey')

    success = {}
    error = {}
    keyerror = {}

    if secret_key == 'secret_123':
        logging.debug("CAUTION!!! MArking result invalid")
        db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

        cursor = db.cursor()
        if not valid:
            success['message'] = "The originIDs [" + originIDs +"] were marked invalid successfully"
            INVALID_ORIGINID_QUERY = "UPDATE result r SET r.isvalid=0 where r.origin_originID in (" + originIDs + ");"
        else:
            success['message'] = "The originIDs [" + originIDs +"] were marked valid successfully"
            INVALID_ORIGINID_QUERY = "UPDATE result r SET r.isvalid=1 where r.origin_originID in (" + originIDs + ");"

        logging.debug(INVALID_ORIGINID_QUERY)
        cursor.execute(INVALID_ORIGINID_QUERY)
        cursor.close()
        db.commit()
        db.close()

    else:
        keyerror['message'] = "BOOM! Wrong Password. This incident will be reported."


    session['success'] = success
    session['error'] = error
    session['keyerror'] = keyerror

    # code = 307 for keeping the original request type ('POST')
    return redirect(url_for('all_runs_secret_page', testname=testname), code=307)

@app.route('/edit-notes', methods=['POST'])
def edit_notes():
    data = json.loads(request.form.get('data'))
    logging.debug(" = {}".format(data))

    originID = data.get('originID')
    testname = data.get('testname')
    new_note = data.get('newNote')

    logging.debug("OriginID = {}\ntestname = {}\nNew Note = {}".format(originID, testname, new_note))

    success = {}
    error = {}
    keyerror = {}

    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                        passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    cursor = db.cursor()
    success['message'] = "The 'notes' of originID '" + str(originID) + "' was changed to '" + new_note + "'  successfully"
    EDIT_NOTES_QUERY = "UPDATE origin SET notes = \'" + new_note + "\' where originID = " + str(originID) + ";"

    logging.debug(EDIT_NOTES_QUERY)
    cursor.execute(EDIT_NOTES_QUERY)
    cursor.close()
    db.commit()
    db.close()

    session['success'] = success
    session['error'] = error
    session['keyerror'] = keyerror

    # code = 307 for keeping the original request type ('POST')
    return redirect(url_for('all_runs_secret_page', testname=testname), code=307)

# Get details for test with originID = 'originID' from database
def get_test_details_data(originID, secret=False):
    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    # Just get the TEST name
    test_name = get_test_name(originID)

    if secret == True:
        RESULTS_VALIDITY_CONDITION = " "
    else:
        RESULTS_VALIDITY_CONDITION = " AND R.isvalid = 1 "

    # Read the subtests description from the wiki_description
    config_file = "./config/wiki_description.ini"
    config_options = configparser.ConfigParser()
    config_options.read(config_file)

    if config_options.has_section(test_name):
        description_string = config_options[test_name]['description'].replace(
            '\"', '')
    else:
        description_string = 'Description'

    logging.debug("BEFORE GETTING DATAFRAME")

    # RESULTS TABLE
    RESULTS_QUERY = """SELECT R.resultID, S.description, R.number, S.resultype, disp.unit, disp.qualifier, R.isvalid FROM result R INNER JOIN subtest S ON S.subtestID=R.subtest_subtestID INNER JOIN display disp ON disp.displayID=R.display_displayID INNER JOIN origin O ON O.originID=R.origin_originID WHERE O.originID=""" + originID + \
                        RESULTS_VALIDITY_CONDITION + ";"
    results_dataframe = pd.read_sql(RESULTS_QUERY, db)

    # Map the resultype to the result type name Example 2-> Single Socket
    index = list(results_dataframe.columns).index('number')
    results_dataframe.insert(index, 'Result Type', [result_type_map.get(result_type, "Unkown resultype") for result_type in results_dataframe['resultype']])

    # Drop the resultype column as it is no longer needed
    del results_dataframe['resultype']

    logging.debug("GOT RESULTS DATAFRAME")
    # logging.debug("{}".format(results_dataframe))

    for col in reversed(description_string.split(',')):
        results_dataframe.insert(1, col, 'default value')

    # Function which splits the description string into various parameters 
    # according to 'description' field of the '.ini' file
    def split_description(index, description):
        try:
            return description.split(',')[index]
        except Exception as error_message:
            logging.debug("Error = {}".format(error_message))
            return np.nan

    # For all the rows in the dataframe, set the description_list values
    description_list = description_string.split(',')
    for j in range(len(description_list)):
        results_dataframe[description_list[j]] = results_dataframe['description'].apply(lambda x: split_description(j, x))

    # Drop the 'description' column as we have now split it into various columns according to description_string
    del results_dataframe['description']

    results_dataframe.dropna(inplace=True)

    # For secret page, return only table data. The secret function will redirect to secret page
    if secret == True:
        secret_context = {
            'testname': test_name,
            'description_list': description_string.split(','),
            'results': results_dataframe.to_dict(orient='list'),
            'originID': originID,
        }

        # close the database connection
        try:
            db.close()
        except:
            pass

        return secret_context
    # Else render test-details.html
    else:
        logging.debug("SECRET WAS FALSE")
        del results_dataframe['resultID']
        del results_dataframe['isvalid']

        # Get some System details
        SYSTEM_DETAILS_QUERY = """SELECT DISTINCT O.hostname, O.testdate, O.notes, O.originID as 'Environment Details',  S.resultype 
                                FROM result R INNER JOIN subtest S ON S.subtestID=R.subtest_subtestID 
                                INNER JOIN origin O ON O.originID=R.origin_originID 
                                WHERE O.originID=""" + originID + ";"
        system_details_dataframe = pd.read_sql(SYSTEM_DETAILS_QUERY, db)

        # Update the Result type (E.g. 0->single thread)
        try:
            system_details_dataframe.update(pd.DataFrame(
                {'resultype': [result_type_map[system_details_dataframe['resultype'][0]]]}))

            # Get result_type and remove it from system_details_dataframe (pop)
            result_type = system_details_dataframe.pop('resultype')[0]
        except:
            result_type = None
            logging.warning('Couldn\'t get Result Type')

        logging.debug(" = {}".format(result_type))

        # Get Num_CPUs list if result_type is 'perf'
        #if result_type == "perf":
        #logging.debug('NUM CPUS START')
        #try:
            # Calls unique_list function on list of unique 'Num_CPUs'
            #num_cpus_list = unique_list((results_dataframe['Num_CPUs']), reverse=True)
        #    raw_dir = '/mnt/nas/dbresults/' + jenkins_details['jobname'][0] + "/" + str(jenkins_details['runID'][0]) + '/results'
        #    logging.debug(raw_dir)
        #    raw_num_cpus_list = [d for d in os.listdir(raw_dir) if os.path.isdir(os.path.join(raw_dir, d))]
        #    num_cpus_list = []
        #    for one_by_one in raw_num_cpus_list:
        #        if one_by_one.isdigit() is True:
        #            num_cpus_list.append(one_by_one)
        #    logging.debug("GOT NUM CPUS")
        #    logging.debug(" = {}".format(num_cpus_list))
        #except Exception as e:
        #    num_cpus_list = []
        #    logging.debug(" = {}".format(e))
            # logging.debug(results_dataframe)
        #    logging.debug("DIDNT GET NUM CPUS")
        #else:
        #    num_cpus_list = []

        logging.debug("Result Type = {}".format(result_type))
        system_details_dataframe = system_details_dataframe.head(1)

        # Get the rest of the system details from jenkins table
        JENKINS_QUERY = """SELECT J.jobname, J.runID FROM origin O INNER JOIN jenkins J 
                            ON O.jenkins_jenkinsID=J.jenkinsID AND O.originID=""" + originID + ";"
        jenkins_details = pd.read_sql(JENKINS_QUERY, db)

        logging.debug('NUM CPUS START')
        num_cpus_list = []
        try:
            # Calls unique_list function on list of unique 'Num_CPUs'
            #num_cpus_list = unique_list((results_dataframe['Num_CPUs']), reverse=True)
            raw_dir = '/mnt/nas/dbresults/' + jenkins_details['jobname'][0] + "/" + str(jenkins_details['runID'][0]) + '/results'
            logging.debug(raw_dir)
            raw_num_cpus_list = [d for d in os.listdir(raw_dir) if os.path.isdir(os.path.join(raw_dir, d))]
            num_cpus_list = []
            for one_by_one in raw_num_cpus_list:
                if one_by_one.isdigit() is True:
                    num_cpus_list.append(one_by_one)
            logging.debug("GOT NUM CPUS")
            logging.debug(" = {}".format(num_cpus_list))
        except Exception as e:
            num_cpus_list = []
            logging.debug(" = {}".format(e))
            # logging.debug(results_dataframe)
            logging.debug("DIDNT GET NUM CPUS")
        #else:
        #    num_cpus_list = []


        # list for creating a column in the system_details_dataframe
        nas_link = []
        nas_link.append("http://localhost:8005/dbresults/" +
                        jenkins_details['jobname'][0] + "/" + str(jenkins_details['runID'][0]))
        jenkins_link = []
        jenkins_link.append("http://localhost:8005/view/Production_Pipeline/job/" +
                            jenkins_details['jobname'][0] + "/" + str(jenkins_details['runID'][0]))

        # Put the Jenkins details in the System Details Dataframe
        system_details_dataframe['NAS Link'] = nas_link
        system_details_dataframe['Jenkins Link'] = jenkins_link
        # SYSTEM details is now ready

        # Check if ramstat.csv exists
        # Ram Utilization Graphs
        nas_path = "/mnt/nas/dbresults/" + jenkins_details['jobname'][0] + '/' + str(jenkins_details['runID'][0]) + '/results/'
        # Get list of directories in '/results/' directory
        dir_list = []
        for x in os.walk(nas_path):
            dir_list = x[1]
            break

        # Get only numeric directories (corresponding to numCPUs)
        dir_list = [x for x in dir_list if x.isnumeric()]

        try:
            ram_file = nas_path + '/' + dir_list[0] + '/ramstat.csv'
        except:
            ram_file = ''
        # Check if ramstat.csv file exists
        if os.path.isfile(ram_file):
            ramstat_csv_exists = True
        else:
            ramstat_csv_exists = False

        try:
            freq_dump_file = nas_path + '/' + dir_list[0] + '/freq_dump.csv'
        except:
            freq_dump_file = ''
        # Check if freq_dump.csv exists
        if os.path.isfile(freq_dump_file):
            freq_dump_csv_exists = True
        else:
            freq_dump_csv_exists = False

        try:
            iostat_csv_file = nas_path +  '/' + dir_list[0] + '/iostat.csv'
        except:
            iostat_csv_file = ''
        # Check if iostat.csv exists
        if os.path.isfile(iostat_csv_file):
            iostat_csv_exists = True
        else:
            iostat_csv_exists = False

        context = {
            'testname': test_name,
            'system_details': system_details_dataframe.to_dict(orient='list'),
            'description_list': description_string.split(','),
            'results': results_dataframe.to_dict(orient='list'),
            'originID': originID,
            
            # Used For 'perf' scaling result_type
            'result_type': result_type,
            'jenkins_details' : {
                'jobname' : jenkins_details['jobname'][0],
                'runID' : str(jenkins_details['runID'][0]),
            },
            'num_cpus_list' : num_cpus_list,
            'ramstat_csv_exists' : ramstat_csv_exists,
            'freq_dump_csv_exists' : freq_dump_csv_exists,
            'iostat_csv_exists' : iostat_csv_exists,
        }

        # close the database connection
        try:
            logging.debug("CLOSING CONNECTION FOR OriginID = {}".format(originID))
            db.close()
        except:
            pass

        return context

# View for handling Test details request
@app.route('/test/<originID>', methods=['GET'])
def test_details_page_old(originID):
    return redirect('/test-details/' + originID)

@app.route('/test-details/<originID>', methods=['GET'])
def test_details_page(originID):
    logging.debug("INSIDE TEST DETAILS FUNCTION = {}".format(originID))
    try:
        context = get_test_details_data(originID)
        error = None
    except Exception as error_message:
        logging.debug("Printing error == {}".format(error_message))
        context = None
        error = error_message

    # For 'Go To Benchmark' Dropdown
    all_tests_data = get_all_tests_data()

    logging.debug("PRINTING CONTEXT = {}".format(context))

    return render_template('test-details.html', error=error, context=context, all_tests_data=all_tests_data)

# Page for marking Individual test 'result' as invalid 
@app.route('/test-details/secret/<originID>', methods=['GET', 'POST'])
def test_details_secret_page(originID):
    if request.method == 'GET':
        # For 'Go To Benchmark' Dropdown
        all_tests_data = get_all_tests_data()

        return render_template('secret-test-details.html', originID={'ID':originID}, context={}, all_tests_data=all_tests_data)
    else:
        success = {}
        error = {}
        keyerror = {}
        logging.debug("#######POSTED###########")
        logging.debug(' = {}'.format(request.args))
        
        # Get doesn't throw error. 
        # If key is not present it sets to default ('None' most of the times)
        success = session.get('success')
        error = session.get('error')
        keyerror = session.get('keyerror')
    
        logging.debug('success = {}'.format(success))
        logging.debug('error = {}'.format(error))
        logging.debug('keyerror = {}'.format(keyerror))

        logging.debug("#######SESSION BEFORE ####")
        logging.debug(' = {}'.format(session))
        logging.debug("########SESSION AFTER $$$$")
        # Clear the contents of the session (cookies)
        session.clear()
        logging.debug(' = {}'.format(session))

        context = get_test_details_data(originID, secret=True)

        # For 'Go To Benchmark' Dropdown
        all_tests_data = get_all_tests_data()

        return render_template('secret-test-details.html', success=success, error=error, keyerror=keyerror, context=context, all_tests_data=all_tests_data)

# Marks a single 'result' invalid
@app.route('/mark-result-id-invalid', methods=['POST'])
def mark_resultID_invalid():
    logging.debug("\n\n\n#REQUEST#########")
    logging.debug(' = {}'.format(request.form))

    data = json.loads(request.form.get('data'))

    logging.debug('Data = {}'.format(data))

    originID = data.get('originID')
    resultIDs = data.get('resultIDs')
    valid = data.get('valid')
    secret_key = data.get('secretKey')

    success = {}
    error = {}
    keyerror = {}

    if secret_key == 'secret_123':
        logging.debug("CAUTION!!! Marking resultID invalid")
        db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

        cursor = db.cursor()
        if not valid:
            success['message'] = """The resultIDs [""" + resultIDs +"""] were marked invalid successfully"""
            CHANGE_RESULTID_VALIDITY_QUERY = "UPDATE result r SET r.isvalid=0 where r.resultID in (" + resultIDs + ");"
        else:
            success['message'] = """The resultIDs [""" + resultIDs +"""] were marked valid successfully"""
            CHANGE_RESULTID_VALIDITY_QUERY = "UPDATE result r SET r.isvalid=1 where r.resultID in (" + resultIDs + ");"

        logging.debug(CHANGE_RESULTID_VALIDITY_QUERY)
        cursor.execute(CHANGE_RESULTID_VALIDITY_QUERY)
        cursor.close()
        db.commit()
        db.close()

    else:
        keyerror['message'] = "BOOM! Wrong Password. This incident will be reported."


    session['success'] = success
    session['error'] = error
    session['keyerror'] = keyerror

    # code = 307 for keeping the original request type ('POST')
    return redirect(url_for('test_details_secret_page', originID=originID), code=307)

# View for handling Environment details request
@app.route('/details/<originID>')
def environment_details_page_old(originID):
    return redirect('/environment-details/' + originID)

@app.route('/environment-details/<originID>')
def environment_details_page(originID):
    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)
    HWDETAILS_QUERY = """SELECT H.fwversion, H.bmcversion, H.biosversion FROM hwdetails H INNER JOIN origin O 
                         ON O.hwdetails_hwdetailsID = H.hwdetailsID AND O.originID = """ + originID + ";"
    hwdetails_dataframe = pd.read_sql(HWDETAILS_QUERY, db)

    TOOLCHAIN_QUERY = """SELECT T.toolchainname, T.toolchainversion, T.flags FROM toolchain as T INNER JOIN origin O
                                 ON O.toolchain_toolchainID = T.toolchainID AND O.originID = """ + originID + ";"
    toolchain_dataframe = pd.read_sql(TOOLCHAIN_QUERY, db)

    OSTUNINGS_QUERY = """SELECT OS.osdistro, OS.osversion, OS.kernelname, OS.pagesize, OS.thp FROM ostunings OS 
                                 INNER JOIN origin O ON O.ostunings_ostuningsID = OS.ostuningsID 
                                 AND O.originID = """ + originID + ";"
    ostunings_dataframe = pd.read_sql(OSTUNINGS_QUERY, db)

    NODE_QUERY = """SELECT N.numsockets, N.skuidname, N.cpuver, N.cpu0serial FROM node as N INNER JOIN hwdetails as H
                            INNER JOIN origin O ON O.hwdetails_hwdetailsID = H.hwdetailsID AND N.nodeID = H.node_nodeID 
                            AND O.originID = """ + originID + ";"
    node_dataframe = pd.read_sql(NODE_QUERY, db)

    BOOTENV_QUERY = """SELECT bootenv.corefreq, bootenv.ddrfreq, bootenv.memnetfreq, bootenv.smt, bootenv.turbo, 
                               bootenv.cores, bootenv.tdp,bootenv.corefeaturemask FROM bootenv INNER JOIN hwdetails H 
                               INNER JOIN origin O ON O.hwdetails_hwdetailsID = H.hwdetailsID 
                               AND H.bootenv_bootenvID = bootenv.bootenvID AND O.originID = """ + originID + ";"
    bootenv_dataframe = pd.read_sql(BOOTENV_QUERY, db)

    JENKINS_QUERY = """SELECT J.jobname, J.runID FROM origin O INNER JOIN jenkins J ON O.jenkins_jenkinsID=J.jenkinsID 
                        AND O.originID = """ + originID + ";"
    jenkins_dataframe = pd.read_sql(JENKINS_QUERY, db)

    #Get the jobname and runID from jenkins_dataframe
    jobname = jenkins_dataframe['jobname'][0]
    runID = jenkins_dataframe['runID'][0]

    # Read RAM, disk and nic details from csv files
    parameter_lists = OrderedDict({
        'ram_details_param_list': [],
        'nic_details_param_list': [],
        'disk_details_param_list': [],

    })

    test_name = get_test_name(originID)

    parameter_lists = read_all_parameter_lists(parameter_lists, test_name)

    results_file_path = '/mnt/nas/dbresults/' + str(jobname) + '/' + str(runID);

    # READ RAM.CSV file and add the size to get total RAM in GB
    try:
        ram_dataframe = pd.read_csv(results_file_path + '/ram.csv', header=None,
                                names=parameter_lists['ram_details_param_list'])

        #CUSTOM FILTER
        # Returns boolean True if the group has all 'ramsize' entries as 'float'
        def only_numeric_groups(df):
            ramsize_series = df['ramsize'].astype(str).str.isnumeric().isin([True])
            return ramsize_series.all()


        ram_dataframe = ram_dataframe.groupby(by=parameter_lists['ram_details_param_list'][0:2]).filter(only_numeric_groups).reset_index(drop=True)

        # Convert each entry of 'ramsize' column to float
        ram_dataframe['ramsize'] = ram_dataframe['ramsize'].apply(lambda x: float(x))

        ram_dataframe = ram_dataframe.groupby(by=parameter_lists['ram_details_param_list'][0:2]) \
                            .apply(lambda x: x.sum()/1024).reset_index()

        #Add ' GB' to the size
        try:
            ram_dataframe['ramsize'] = ram_dataframe['ramsize'].apply(lambda x: str(x) + " GB")
        except:
            logging.warning("Couldn't add 'GB' to RAM SIZE")
            pass
    except Exception as e:
        logging.debug("Error: ", e)
        ram_dataframe = pd.DataFrame()

    # Read disk dataframe
    try:
        disk_dataframe = pd.read_csv(results_file_path + '/disk.csv', header=None,
                                 names=parameter_lists['disk_details_param_list'])
    except Exception as e:
        logging.debug("Error: ", e)
        disk_dataframe = pd.DataFrame()
    
    try:
        nic_dataframe = pd.read_csv(results_file_path + '/nic.csv', header=None,
                                names=parameter_lists['nic_details_param_list'])
    except Exception as e:
        logging.debug("Error: ", e)
        nic_dataframe = pd.DataFrame()


    context = {
        'originID' : originID,
        'hwdetails': hwdetails_dataframe.to_dict(orient='list'),
        'toolchain': toolchain_dataframe.to_dict(orient='list'),
        'ostunings': ostunings_dataframe.to_dict(orient='list'),
        'node':      node_dataframe.to_dict(orient='list'),
        'bootenv':   bootenv_dataframe.to_dict(orient='list'),
        'ram':       ram_dataframe.to_dict(orient='list'),
        'disk':      disk_dataframe.to_dict(orient='list'),
        'nic':       nic_dataframe.to_dict(orient='list'),
    }

    # close the database connection
    try:
        db.close()
    except:
        pass

    # For 'Go To Benchmark' Dropdown
    all_tests_data = get_all_tests_data()

    return render_template('environment-details.html', context=context, all_tests_data=all_tests_data)

# Compare two or more tests
@app.route('/diff', methods=['GET'])
def diff_tests():
    start_time = time.time()

    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                            passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    logging.debug(request)
    logging.debug("REQUEST ARGS = {}".format(request.args))

    # take checked rows from table
    originID_compare_list = [value for key, value in request.args.items() if "diff-checkbox" in key]

    originID_compare_list.sort()
    logging.debug(' = {}'.format(originID_compare_list))

    # Get the Test name
    test_name = get_test_name(originID_compare_list[0])

    JENKINS_QUERY = """SELECT J.jobname, J.runID FROM origin O INNER JOIN jenkins J 
                        ON O.jenkins_jenkinsID=J.jenkinsID AND O.originID in """ \
                        + str(originID_compare_list).replace('[', '(').replace(']', ')') + ";"
    jenkins_details = pd.read_sql(JENKINS_QUERY, db)

    #Get the jobname and runID from jenkins details
    jobname_list = jenkins_details['jobname'].to_list()
    runID_list = jenkins_details['runID'].to_list()

    # create all parameter lists
    parameter_lists = OrderedDict({
        'results_param_list': [],
        'origin_param_list': [],
        'bootenv_param_list': [],
        'node_param_list': [],
        'hwdetails_param_list': [],
        'ostunings_param_list': [],
        'toolchain_param_list': [],
        'ram_details_param_list': [],
        'nic_details_param_list': [],
        'disk_details_param_list': [],

        'qualifier': [],
        'min_or_max': [],
    })

    parameter_lists = read_all_parameter_lists(parameter_lists, test_name)

    # removes 'qualifier' from parameter_lists and assigns to qualifier_list variable (a list)
    qualifier_list = parameter_lists.pop('qualifier')
    min_or_max_list = parameter_lists.pop('min_or_max')

    # Dictionary of List of Dictionaries (tests) to be compared
    compare_lists = OrderedDict({
        'origin_list': [],
        'bootenv_list': [],
        'node_list': [],
        'hwdetails_list': [],
        'ostunings_list': [],
        'toolchain_list': [],
        'ram_details_list': [],
        'nic_details_list': [],
        'disk_details_list': [],
    })

    # Function for reading "results.csv". 
    # For Improved readability
    def read_results():
        # Read all results. Store in a dataframe
        join_on_columns_list = parameter_lists['results_param_list'][0:-4]
        join_on_columns_list.extend(['resultype', 'unit', 'qualifier'])

        # read the first results file
        results_file_path = '/mnt/nas/dbresults/' + str(jobname_list[0]) + '/' \
                            + str(runID_list[0]) + '/results/results.csv'
        first_results_dataframe = pd.read_csv(results_file_path, header=None,
                                                names=parameter_lists['results_param_list'])

        # All columns as string type^M
        first_results_dataframe[join_on_columns_list] = first_results_dataframe[join_on_columns_list].astype(str)
        # LOWERCASE THE QUALIFIER COLUMN
        first_results_dataframe['qualifier'] = first_results_dataframe['qualifier'].apply(lambda x: x.lower().strip())

        # GROUP BY join_on_columns_list AND FIND MIN/MAX OF EACH GROUP
        if min_or_max_list[0] == '0':
            first_results_dataframe = first_results_dataframe.groupby(by=join_on_columns_list).min()
        else:
            first_results_dataframe = first_results_dataframe.groupby(by=join_on_columns_list).max()

        # Assign first results_dataframe to intermediate_dataframe for merging
        intermediate_dataframe = first_results_dataframe

        logging.debug('\n\nDONE\n\n')

        # for each subsequent results file, merge with the already exsiting dataframe on "description" columns
        for jobname, runID in zip(jobname_list[1:], runID_list[1:]):
            results_file_path = '/mnt/nas/dbresults/' + str(jobname) + '/' + str(runID) + '/results/results.csv'

            next_dataframe = pd.read_csv(results_file_path, header=None, names=parameter_lists['results_param_list'])
            # All columns as string type
            next_dataframe[join_on_columns_list] = next_dataframe[join_on_columns_list].astype(str)
            next_dataframe['qualifier'] = next_dataframe['qualifier'].apply(lambda x: x.lower().strip())

            # GROUP BY join_on_columns_list AND FIND MIN/MAX OF EACH GROUP
            if min_or_max_list[0] == '0':
                next_dataframe = next_dataframe.groupby(by=join_on_columns_list).min()
            else:
                next_dataframe = next_dataframe.groupby(by=join_on_columns_list).max()

            # Merge the next_dataframe with previous
            intermediate_dataframe = intermediate_dataframe.merge(next_dataframe, how='outer', on=join_on_columns_list,
                                                                    validate="many_to_many")

        # Change column names according to OriginID
        intermediate_dataframe = intermediate_dataframe.reset_index()

        intermediate_dataframe.columns = join_on_columns_list + ["number_" + originID for originID in
                                                                    originID_compare_list]

        def apply_result_type(x):
            try:
                return result_type_map[int(x)]
            except:
                logging.warning("Couldn't parse result type")
                return None

        intermediate_dataframe['resultype'] = intermediate_dataframe['resultype'].apply(lambda x: apply_result_type(x))

        final_results_dataframe = intermediate_dataframe.sort_values(by=['qualifier'])
        logging.debug("PRINTING THE FINAL DATAFRAME")
        logging.debug(' = {}'.format(final_results_dataframe))
        logging.debug("DONE")

        # DROP THE NAN rows for comparing results in graphs
        comparable_results = final_results_dataframe.dropna()
        comparable_results = comparable_results[
            ['qualifier'] + [column for column in comparable_results.columns if column not in join_on_columns_list]]

        comparable_results.columns = ['qualifier'] + ["Test_" + originID for originID in originID_compare_list]

        # FILL NAN cells with ""
        final_results_dataframe = final_results_dataframe.fillna("")

        logging.debug("PRINTING COMPARABLE RESULTS")
        logging.debug(' = {}'.format(comparable_results))
        logging.debug("DONE")

        return final_results_dataframe, comparable_results, join_on_columns_list

    # Get Results table, comparable_results(unused),
    # Join on columns list is (description_list + resultype + unit + qualifier)
    final_results_dataframe, comparable_results, join_on_columns_list = read_results()

    # Function for reading ram_dataframe
    def read_ram_details():
        # READ RAM.CSV file and add the size to get total RAM in GB
        ram_dataframe = pd.read_csv(results_file_path + '/ram.csv', header=None,
                                    names=parameter_lists['ram_details_param_list'])

        #CUSTOM FILTER
        # Returns boolean True if the group has all 'ramsize' entries as 'float'
        def only_numeric_groups(df):
            ramsize_series = df['ramsize'].astype(str).str.isnumeric().isin([True])
            return ramsize_series.all()


        ram_dataframe = ram_dataframe.groupby(by=parameter_lists['ram_details_param_list'][0:2]).filter(only_numeric_groups).reset_index(drop=True)

        # Convert each entry of 'ramsize' column to float
        ram_dataframe['ramsize'] = ram_dataframe['ramsize'].apply(lambda x: float(x))

        ram_dataframe = ram_dataframe.groupby(by=parameter_lists['ram_details_param_list'][0:2]) \
                            .apply(lambda x: x.sum()/1024).reset_index()

        #Add ' GB' to the size
        try:
            ram_dataframe['ramsize'] = ram_dataframe['ramsize'].apply(lambda x: str(x) + " GB")
        except:
            logging.warning("Couldn't add 'GB' to RAM SIZE")
            pass

        return ram_dataframe
        pass

    def read_disk_details():
        pass

    def read_nic_details():
        pass

    # Delete the param_lists which are no longer needed
    del parameter_lists['results_param_list']
    # del parameter_lists['ram_details_param_list'],
    # del parameter_lists['nic_details_param_list'],
    # del parameter_lists['disk_details_param_list'],

    # read csv files from NAS path
    compare_lists = read_all_csv_files(compare_lists, parameter_lists, originID_compare_list)

    # send data to the template compare.html
    context = {
        'originID_list': originID_compare_list,
        'testname': test_name,

        'index_columns': join_on_columns_list,

        'parameter_lists': parameter_lists,
        'compare_lists': compare_lists,

        'comparable_results': comparable_results.to_dict(orient='list'), #Unused as of now
        'results': final_results_dataframe.to_dict(orient='list'),
    }
    # close the database connection
    try:
        db.close()
    except:
        pass

    # For 'Go To Benchmark' Dropdown
    all_tests_data = get_all_tests_data()

    return render_template('compare.html', context=context, all_tests_data=all_tests_data)

# This function handles the AJAX request for Comparison graph data. 
# JS then draws the graph using this data
@app.route('/sku_comparison_graph', methods=['POST'])
def sku_comparison_graph():
    logging.debug("Got the request for Sku Comparison Graph")

    start_time = time.time()

    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    data = request.get_json()
    xParameter = data['xParameter']
    yParameter = data['yParameter']

    testname = data['testname']

    # Result type filter eg. dual socket
    result_type_filter = data['resultTypeFilter']

    if result_type_filter == "None":
        result_type_filter = None

    # Get input_filter_condition by calling the function
    input_filters_list = data['inputFiltersList']
    INPUT_FILTER_CONDITION = get_input_filter_condition(testname, input_filters_list)

    results_metadata_file_path = './config/wiki_description.ini'
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(results_metadata_file_path)

    # GET qualifier_list and min_max_list from 'fields' and 'higher_is_better' the section 'testname'
    qualifier_list = results_metadata_parser.get(testname, 'fields').replace('\"', '').split(',')
    min_or_max_list = results_metadata_parser.get(testname, 'higher_is_better') \
                    .replace('\"', '').replace(' ', '').split(',')
    index = qualifier_list.index(yParameter)

    sku_file_path = './config/sku_definition.ini'
    sku_parser = configparser.ConfigParser()
    sku_parser.read(sku_file_path)

    # Dictionary mapping from 'skuidname' : 'server_cpu_name'
    # Example 'Cavium ThunderX2(R) CPU CN9980 v2.2 @ 2.20 GHz' : Marvell TX2-B2
    skuid_cpu_map = OrderedDict({section: sku_parser.get(section, 'SKUID').replace('\"', '').split(',') for section in sku_parser.sections()})

    # Fill the sku_cpu_map with all "sku->section" mapping entries
    for section in sku_parser.sections():
        skus = sku_parser.get(section, 'SKUID').replace('\"','').split(',')
        for sku in skus:
            skuid_cpu_map[sku] = section


    parameter_map = {
        "Kernel Version": 'os.kernelname', 
        'OS Version': 'os.osversion', 
        'OS Name': 'os.osdistro', 
        "Firmware Version": 'hw.fwversion' , 
        "ToolChain Name": 'tc.toolchainname', 
        "ToolChain Version" : 'tc.toolchainversion', 
        "Flags": 'tc.flags',
        "SMT" : 'b.smt',
        "Cores": 'b.cores',
        "Corefreq": 'b.corefreq',
        "DDRfreq": 'b.ddrfreq',
        "SKUID": 'n.skuidname',
        "Hostname": 'o.hostname',
        "Scaling" : 's.resultype',
    }
    join_on_map_for_x_query = {
        'Kernel Version': 'INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID',
        'OS Version': 'INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID',
        'OS Name': 'INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID',
        "Firmware Version": 'INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID', 
        "ToolChain Name": 'INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID',
        "ToolChain Version" : 'INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID',
        "Flags": 'INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID',
        "SMT" : 'INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "Cores": 'INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "Corefreq": 'INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "DDRfreq": 'INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "SKUID": 'INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID INNER JOIN node n ON hw.node_nodeID = n.nodeID',
        "Hostname" : ' ',
        "Scaling" : ' ',
    }
    join_on_map = {
        'Kernel Version': 'INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID',
        'OS Version': 'INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID',
        'OS Name': 'INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID',
        "Firmware Version": ' ',
        "ToolChain Name": 'INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID',
        "ToolChain Version" : 'INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID',
        "Flags": 'INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID',
        "SMT" : 'INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "Cores": 'INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "Corefreq": 'INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "DDRfreq": 'INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "SKUID": ' ',
        "Hostname" : ' ',
        "Scaling" : ' ',
    }

    # Conditions for filtering x_list
    filter_x_list_map = {
        "Kernel Version": 'not empty string', 
        'OS Version': 'not empty string',
        'OS Name': 'not empty string',
        "Firmware Version": 'not empty string',
        "ToolChain Name": 'not empty string',
        "ToolChain Version" : 'not empty string',
        "Flags": 'not empty string',
        "SMT" : 'greater than zero',
        "Cores": 'greater than zero',
        "Corefreq": 'greater than zero',
        "DDRfreq": 'greater than zero',
        "SKUID": 'not empty string',
        "Hostname": 'not empty string',
        "Scaling" : '',
    }

    server_cpu_list = []

    # List of lists
    # Each list has entries for a single CPU Manufacturer
    x_list_list = []
    y_list_list = []
    originID_list_list = []

    if xParameter == "Scaling":
        # [dual socket, single socket, 1/2 socket, 1/4th socket, 1/8th  socket, 2 cores, single core, single thread]
        initial_x_list = [3, 2, 7, 6, 5, 8, 1, 0]
    else:
        # Get initial_x_list by excecuting the query
        X_LIST_QUERY = "SELECT DISTINCT " + parameter_map[xParameter] + " as \'" + parameter_map[xParameter] + \
                        """\' FROM origin o INNER JOIN result r ON o.originID = r.origin_originID 
                        INNER JOIN testdescriptor t ON t.testdescriptorID=o.testdescriptor_testdescriptorID
                        INNER JOIN subtest s ON r.subtest_subtestID=s.subtestID 
                        INNER JOIN display disp ON  r.display_displayID = disp.displayID """ + \
                        join_on_map_for_x_query[xParameter] + \
                        " WHERE t.testname=\'""" + testname + "\'" + INPUT_FILTER_CONDITION + ";"

        logging.debug("Printing XLIST QUERY")
        logging.debug(X_LIST_QUERY)

        x_df = pd.read_sql(X_LIST_QUERY, db)
        logging.debug("X_Dataframe = {}".format(x_df))
        initial_x_list = sorted(x_df[parameter_map[xParameter]].to_list())
        logging.debug("X_LIST = {}".format(initial_x_list))

    if filter_x_list_map[xParameter] == 'not empty string':
        # Convert each element to type "str"
        initial_x_list = list(map(lambda x: str(x).strip(), initial_x_list))
        
        # Remove ALL the entries which are '' in the list 
        initial_x_list = list(filter(lambda x: x != '', initial_x_list))
    elif filter_x_list_map[xParameter] == 'greater than zero':
        initial_x_list = list(filter(lambda x: x > 0, initial_x_list))
    else:
        pass

    logging.debug("initial_x_list = {}".format(initial_x_list))

    if min_or_max_list[index] == '0':
        RESULTS_QUERY = "SELECT r.number, o.originID, n.skuidname as 'skuidname_legend', s.resultype as 'resultype_filter', r.isvalid, " + \
                        parameter_map[xParameter] + """ FROM origin o 
                        INNER JOIN result r ON o.originID = r.origin_originID 
                        INNER JOIN display disp ON  r.display_displayID = disp.displayID 
                        INNER JOIN testdescriptor t ON t.testdescriptorID = o.testdescriptor_testdescriptorID 
                        INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID 
                        INNER JOIN node n ON hw.node_nodeID = n.nodeID 
                        INNER JOIN subtest s ON r.subtest_subtestID=s.subtestID """ + join_on_map[xParameter] + \
                        " WHERE t.testname = \'" + testname + "\' AND r.number > 0 " + \
                        " AND disp.qualifier LIKE \'%" + qualifier_list[index] + "%\'" + \
                        INPUT_FILTER_CONDITION + ";"
    else:
        RESULTS_QUERY = "SELECT r.number, o.originID, n.skuidname as 'skuidname_legend', s.resultype as 'resultype_filter', r.isvalid, " + \
                        parameter_map[xParameter] + """ FROM origin o 
                        INNER JOIN result r ON o.originID = r.origin_originID 
                        INNER JOIN display disp ON  r.display_displayID = disp.displayID 
                        INNER JOIN testdescriptor t ON t.testdescriptorID = o.testdescriptor_testdescriptorID 
                        INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID 
                        INNER JOIN node n ON hw.node_nodeID = n.nodeID 
                        INNER JOIN subtest s ON r.subtest_subtestID=s.subtestID """ + join_on_map[xParameter] + \
                      " WHERE t.testname = \'" + testname + \
                        "\' AND disp.qualifier LIKE \'%" + qualifier_list[index] + "%\'" + \
                        INPUT_FILTER_CONDITION + ";"

    results_df = pd.read_sql(RESULTS_QUERY, db)
    logging.debug(RESULTS_QUERY)
    logging.debug(results_df.shape)

    if not results_df.empty:
        # Remove results which are not valid
        results_df = results_df[results_df['isvalid'] == 1].reset_index(drop=True)
        del results_df['isvalid']
        logging.debug(results_df.shape)

        # Convert to actual result type string
        results_df['resultype_filter'] = results_df['resultype_filter'].apply(lambda x : result_type_map.get(x, "Unkown"))
        if result_type_filter:
            results_df = results_df[results_df['resultype_filter'] == result_type_filter].reset_index(drop=True)

        del results_df['resultype_filter']

        # If results_df is not empty after applying reusult_type_filter
        if not results_df.empty:
            results_df['skuidname_legend'] = results_df['skuidname_legend'].apply(lambda x: x.strip())

            # Only skuidnames which are in sku_definition.ini will be shown
            valid_skuidnames = list(itertools.chain(*[sku_parser.get(section, 'SKUID').replace('\"', '').split(',') \
                                for section in sku_parser.sections()]))
            results_df = results_df[results_df['skuidname_legend']
                                .apply(lambda x: x.strip() in valid_skuidnames)].reset_index(drop=True)

            # Filter on initial_x_list
            param_name = parameter_map[xParameter]
            param_name = param_name[param_name.find('.')+1:]

            results_df = results_df[results_df[param_name]
                                .apply(lambda x: str(x).strip().upper() in [str(e).strip() for e in map(str.upper, list(map(str, initial_x_list)))])].reset_index(drop=True)

            # Convert skuidname to corresponding section in sku_definition.ini
            results_df['skuidname_legend'] = results_df['skuidname_legend'].apply(lambda x: skuid_cpu_map[x])

            # Get Min or Max results and group them w.r.t. skuidname_legend and x parameter
            if min_or_max_list[index] == '0':
                idx = results_df.groupby(by=['skuidname_legend', param_name])['number'].idxmin()
            else:
                idx = results_df.groupby(by=['skuidname_legend', param_name])['number'].idxmax()

            results_df = results_df.loc[idx].sort_values(by=[param_name, 'skuidname_legend']).reset_index(drop=True)

            # Sort by skuidname_legend according to the sku_definition.ini file
            sku_categories = sku_parser.sections()
            ordered_skus = pd.Categorical(results_df['skuidname_legend'].tolist(), categories = sku_categories, ordered=True)
            results_df['skuidname_legend'] = pd.Series(ordered_skus)
            results_df = results_df.sort_values(by='skuidname_legend').reset_index(drop=True)

            # Map the result type according to result_type_map
            if xParameter == 'Scaling':
                results_df['resultype'] = results_df['resultype'].apply(lambda x : result_type_map[x])

                # Sort by category
                result_type_categories = [result_type_map[x] for x in initial_x_list]
                ordered_resultypes = pd.Categorical(results_df['resultype'].tolist(), categories=[result_type_map[x] for x in initial_x_list], ordered=True)
                results_df['resultype'] = pd.Series(ordered_resultypes)
                results_df = results_df.sort_values(by=['skuidname_legend', 'resultype']).reset_index(drop=True)

            # Get unique skuidname entries
            server_cpu_list = [x for x in results_df['skuidname_legend'].unique().tolist()]

            results_df = results_df.set_index('skuidname_legend')

            # Extract the lists from the dataframe
            # Always pass a list to .loc function to get Dataframe as the result
            x_list_list = [results_df.loc[[skuidname]][param_name].tolist() for skuidname in server_cpu_list]
            y_list_list = [results_df.loc[[skuidname]]['number'].tolist() for skuidname in server_cpu_list]
            originID_list_list = [results_df.loc[[skuidname]]['originID'].tolist() for skuidname in server_cpu_list]

    # Get colours for cpu manufacturer
    color_list = []
    visibile_list = []
    for section in server_cpu_list:
        color_list.extend(sku_parser.get(section, 'color').replace('\"','').split(','))
        visibile_list.extend(sku_parser.get(section, 'visible').replace('\"','').split(','))

    # Get the unit for the selected yParamter (qualifier)
    UNIT_QUERY = """SELECT disp.qualifier, disp.unit FROM origin o INNER JOIN testdescriptor t 
                    ON t.testdescriptorID=o.testdescriptor_testdescriptorID  INNER JOIN result r 
                    ON o.originID = r.origin_originID  INNER JOIN display disp ON  r.display_displayID = disp.displayID 
                    where t.testname = \'""" + testname + "\' and disp.qualifier LIKE \'%" + yParameter.strip() +"%\' limit 1;"
    unit_df = pd.read_sql(UNIT_QUERY, db)

    logging.debug("Printing UNIT Query")
    logging.debug(UNIT_QUERY)
    try:
        y_axis_unit = unit_df['unit'][0]
    except Exception as error_message:
        y_axis_unit = "Unknown Unit"
        logging.debug("= {}".format(error_message))

    response = {
        'x_list_list': x_list_list, 
        'y_list_list': y_list_list,
        'y_axis_unit': y_axis_unit,
        'xParameter': xParameter,
        'yParameter': yParameter,
        'originID_list_list': originID_list_list,
        'server_cpu_list': server_cpu_list,
        'color_list': color_list,
        'visible_list':visibile_list,
    }

    logging.debug("Printing Final response")
    logging.debug(response)

    # close the database connection
    try:
        logging.debug("CLOSING CONNECTION FOR sku_comparison_graph {}".format(testname))
        db.close()
    except:
        pass

    print("sku_comparison_graph took {} seconds".format(time.time() - start_time))

    return response

# This function handles the AJAX request for Best SKU Graph data.
# JS then draws the graph using this data
@app.route('/best_sku_graph', methods=['POST'])
def best_sku_graph():
    start_time = time.time()

    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    data = request.get_json()
    xParameter = data['xParameter'] # Not used. Sending ""
    yParameter = data['yParameter'] # Not used. Sending "Best" + first qualifier as yParameter
    testname = data['testname']

    # Result type filter eg. dual socket
    result_type_filter = data['resultTypeFilter']

    if result_type_filter == "None":
        result_type_filter = None

    results_metadata_file_path = './config/wiki_description.ini'
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(results_metadata_file_path)

    # For the input filters
    input_parameters = results_metadata_parser.get(testname, 'description') \
                                                .replace('\"', '').replace(' ', '').split(',')
    INPUT_FILTER_CONDITION = ""
    try:
        input_filters_list = data['inputFiltersList']
        
        for index, input_filter in enumerate(input_filters_list):
            if(input_filter != "None"):
                if(input_filter.isnumeric()):
                    INPUT_FILTER_CONDITION += " and SUBSTRING_INDEX(SUBSTRING_INDEX(s.description,','," + str(index+1) +"),',',-1)=" + input_filter 
                else:
                    INPUT_FILTER_CONDITION += " and SUBSTRING_INDEX(SUBSTRING_INDEX(s.description,','," + str(index+1) +"),',',-1)=\'" + input_filter  + "\'"
    except:
        pass

    # GET First qualifier from 'fields' and its corresponding min_or_max from 'higher_is_better' for section 'testname'
    qualifier = results_metadata_parser.get(testname, 'fields').replace('\"', '').split(',')[0]
    min_or_max = results_metadata_parser.get(testname, 'higher_is_better') \
                    .replace('\"', '').replace(' ', '').split(',')[0]


    sku_file_path = './config/sku_definition.ini'
    sku_parser = configparser.ConfigParser()
    sku_parser.read(sku_file_path)

    # Dictionary mapping from 'skuidname' : 'server_cpu_name'
    # Example 'Cavium ThunderX2(R) CPU CN9980 v2.2 @ 2.20 GHz' : Marvell TX2-B2
    skuid_cpu_map = OrderedDict({section: sku_parser.get(section, 'SKUID').replace('\"', '').split(',') for section in sku_parser.sections()})

    # Fill the sku_cpu_map with all "sku->section" mapping entries
    for section in sku_parser.sections():
        skus = sku_parser.get(section, 'SKUID').replace('\"','').split(',')
        for sku in skus:
            skuid_cpu_map[sku] = section

    x_list = []
    y_list = []
    originID_list = []

    if min_or_max == '0':
        BEST_RESULT_QUERY = """SELECT r.number, o.originID, n.skuidname, s.resultype as 'resultype_filter', r.isvalid FROM origin o 
                            INNER JOIN result r on r.origin_originID = o.originID 
                            INNER JOIN display disp ON r.display_displayID = disp.displayID 
                            INNER JOIN testdescriptor t on t.testdescriptorID = o.testdescriptor_testdescriptorID 
                            INNER JOIN hwdetails hw on hw.hwdetailsID = o.hwdetails_hwdetailsID 
                            INNER JOIN node n on n.nodeID = hw.node_nodeID 
                            INNER JOIN subtest s ON r.subtest_subtestID=s.subtestID """ + \
                            " WHERE t.testname = \'" + testname + "\' AND r.number > 0 " + \
                            " AND disp.qualifier LIKE \'%" + qualifier + "%\'" + \
                            INPUT_FILTER_CONDITION + ";"
    else:
        BEST_RESULT_QUERY = """SELECT r.number, o.originID, n.skuidname, s.resultype as 'resultype_filter', r.isvalid FROM origin o 
                            INNER JOIN result r on r.origin_originID = o.originID 
                            INNER JOIN display disp ON r.display_displayID = disp.displayID 
                            INNER JOIN testdescriptor t on t.testdescriptorID = o.testdescriptor_testdescriptorID 
                            INNER JOIN hwdetails hw on hw.hwdetailsID = o.hwdetails_hwdetailsID 
                            INNER JOIN node n on n.nodeID = hw.node_nodeID 
                            INNER JOIN subtest s ON r.subtest_subtestID=s.subtestID """ + \
                            " WHERE t.testname = \'" + testname + "\' " + \
                            " AND disp.qualifier LIKE \'%" + qualifier + "%\'" + \
                            INPUT_FILTER_CONDITION + ";"

    results_df = pd.read_sql(BEST_RESULT_QUERY, db)

    if not results_df.empty:
        # Remove results which are not valid
        results_df = results_df[results_df['isvalid'] == 1].reset_index(drop=True)
        del results_df['isvalid']
        logging.debug(results_df.shape)

        # Convert to actual result type string
        results_df['resultype_filter'] = results_df['resultype_filter'].apply(lambda x : result_type_map.get(x, "Unkown"))
        if result_type_filter:
            results_df = results_df[results_df['resultype_filter'] == result_type_filter].reset_index(drop=True)

        del results_df['resultype_filter']

        # If results_df is not empty after applying reusult_type_filter
        if not results_df.empty:
            # Strip all the skuidnames
            results_df['skuidname'] = results_df['skuidname'].apply(lambda x: x.strip())

            # Only skuidnames which are in sku_definition.ini will be shown
            valid_skuidnames = list(itertools.chain(*[sku_parser.get(section, 'SKUID').replace('\"', '').split(',') \
                                for section in sku_parser.sections()]))
            results_df = results_df[results_df['skuidname']
                                .apply(lambda x: x.strip() in valid_skuidnames)].reset_index(drop=True)

            # Convert skuidname to corresponding section in sku_definition.ini
            results_df['skuidname'] = results_df['skuidname'].apply(lambda x: skuid_cpu_map[x])

            # Get max/min results
            if min_or_max == '0':
                idx = results_df.groupby(by=['skuidname'])['number'].idxmin()
            else:
                idx = results_df.groupby(by=['skuidname'])['number'].idxmax()
            results_df = results_df.loc[idx].reset_index(drop=True)

            # Sort by skuidname according to the sku_definition.ini file
            sku_categories = sku_parser.sections()
            ordered_skus = pd.Categorical(results_df['skuidname'].tolist(), categories = sku_categories, ordered=True)
            results_df['skuidname'] = pd.Series(ordered_skus)
            results_df = results_df.sort_values(by='skuidname')

            x_list = results_df['skuidname'].tolist()
            y_list = results_df['number'].tolist()
            originID_list = results_df['originID'].tolist()

    logging.debug(results_df)

    color_list = []
    for section in x_list:
        color_list.extend(sku_parser.get(section, 'color').replace('\"','').split(','))
    logging.debug("= {}".format(color_list))

    # Get the unit for the selected yParamter (qualifier)
    UNIT_QUERY = """SELECT disp.qualifier, disp.unit FROM origin o INNER JOIN testdescriptor t 
                    ON t.testdescriptorID=o.testdescriptor_testdescriptorID  INNER JOIN result r 
                    ON o.originID = r.origin_originID  INNER JOIN display disp ON  r.display_displayID = disp.displayID 
                    where t.testname = \'""" + testname + "\' and disp.qualifier LIKE \'%" + qualifier +"%\' limit 1;"
    unit_df = pd.read_sql(UNIT_QUERY, db)
    y_axis_unit = unit_df['unit'][0]

    response = {
        'x_list': x_list, 
        'y_list': y_list,
        'y_axis_unit': y_axis_unit,
        'color_list': color_list,
        'xParameter': "",
        'yParameter': 'Best ' + qualifier,
        'originID_list': originID_list,
        'higher_is_better': min_or_max,
    }

    # close the database connection
    try:
        db.close()
    except:
        pass

    logging.debug(response)
    print("best_sku_graph took {} seconds".format(time.time() - start_time))

    return response

# Returns the NORMALIZED version of the graph with respect to a xParameter
@app.route('/best_sku_graph_normalized', methods=['POST'])
def best_sku_graph_normalized():
    results_metadata_file_path = './config/wiki_description.ini'
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(results_metadata_file_path)

    data = request.get_json()

    # X-Axis, Y-Axis values lists and its parameters
    x_list = data['xList']
    y_list = data['yList']
    xParameter = data['xParameter']
    yParameter = data['yParameter']
    originID_list = data['originIDList']
    testname = data['testname']

    # Normalized with respect to this parameter
    normalized_wrt = data['normalizedWRT']

    # Higher is better => 0 or 1
    higher_is_better = results_metadata_parser.get(testname, 'higher_is_better') \
                        .replace('\"', '').replace(' ', '').split(',')[0]

    # If lower is better, then take the inverse of the normalized values
    index = x_list.index(normalized_wrt)
    logging.debug("PRINTING X LIST= {}".format(x_list))
    logging.debug("FOUND AT INDEX = {}".format(index))
    if higher_is_better == "1":
        logging.debug("HIGHER IS BETTER")
        logging.debug(y_list)

        normalized_y_list = [value/y_list[index] for value in y_list]
    else:
        # Inverse
        logging.debug("LOWER IS BETTER")
        logging.debug(y_list)
       # normalized_y_list = [y_list[index]/value for value in y_list]
        normalized_y_list = [value/y_list[index] for value in y_list]

    # Colors for the graphs
    sku_file_path = './config/sku_definition.ini'
    sku_parser = configparser.ConfigParser()
    sku_parser.read(sku_file_path)

    color_list = []
    for section in x_list:
        color_list.extend(sku_parser.get(section, 'color').replace('\"','').split(','))
    logging.debug(" = {}".format(color_list))

    logging.debug("Normalized Y list = {}".format(normalized_y_list))

    response = {
        'x_list': x_list, 
        'y_list': normalized_y_list,
        'y_axis_unit': "ratio",
        'color_list': color_list,
        'xParameter': xParameter,
        'yParameter': yParameter,
        'originID_list': originID_list,
        'higher_is_better': higher_is_better,
    }

    return response

# Timeline Graph
@app.route('/timeline_graph', methods=['POST'])
def timeline_graph():
    logging.debug("Got the request for Timeline Graph")

    start_time = time.time()

    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    data = request.get_json()
    xParameter = data['xParameter']
    yParameter = data['yParameter']
    smt_filter = data['smtFilter']
    sku_filter = data['skuFilter']

    logging.debug("Printing data\n\n{}".format(data))

    testname = data['testname']

    # Result type filter eg. dual socket
    result_type_filter = data['resultTypeFilter']

    if result_type_filter == "None":
        result_type_filter = None

    # Get input_filter_condition by calling the function
    input_filters_list = data['inputFiltersList']
    INPUT_FILTER_CONDITION = get_input_filter_condition(testname, input_filters_list)

    results_metadata_file_path = './config/wiki_description.ini'
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(results_metadata_file_path)

    # GET qualifier_list and min_max_list from 'fields' and 'higher_is_better' the section 'testname'
    qualifier_list = results_metadata_parser.get(testname, 'fields').replace('\"', '').split(',')
    min_or_max_list = results_metadata_parser.get(testname, 'higher_is_better') \
                    .replace('\"', '').replace(' ', '').split(',')
    index = qualifier_list.index(yParameter)

    sku_file_path = './config/sku_definition.ini'
    sku_parser = configparser.ConfigParser()
    sku_parser.read(sku_file_path)

    # Dictionary mapping from 'skuidname' : 'server_cpu_name'
    # Example 'Cavium ThunderX2(R) CPU CN9980 v2.2 @ 2.20 GHz' : Marvell TX2-B2
    skuid_cpu_map = OrderedDict({section: sku_parser.get(section, 'SKUID').replace('\"', '').split(',') for section in sku_parser.sections()})

    # Fill the sku_cpu_map with all "sku->section" mapping entries
    for section in sku_parser.sections():
        skus = sku_parser.get(section, 'SKUID').replace('\"','').split(',')
        for sku in skus:
            skuid_cpu_map[sku] = section


    parameter_map = {
        "Kernel Version": 'os.kernelname', 
        'OS Version': 'os.osversion', 
        'OS Name': 'os.osdistro', 
        "Firmware Version": 'hw.fwversion' , 
        "ToolChain Name": 'tc.toolchainname', 
        "ToolChain Version" : 'tc.toolchainversion', 
        "Flags": 'tc.flags',
        "SMT" : 'b.smt',
        "Cores": 'b.cores',
        "Corefreq": 'b.corefreq',
        "DDRfreq": 'b.ddrfreq',
        "SKUID": 'n.skuidname',
        "Hostname": 'o.hostname',
        "Scaling" : 's.resultype',
    }
    join_on_map_for_x_query = {
        'Kernel Version': 'INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID',
        'OS Version': 'INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID',
        'OS Name': 'INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID',
        "Firmware Version": 'INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID', 
        "ToolChain Name": 'INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID',
        "ToolChain Version" : 'INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID',
        "Flags": 'INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID',
        "SMT" : 'INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "Cores": 'INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "Corefreq": 'INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "DDRfreq": 'INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "SKUID": 'INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID INNER JOIN node n ON hw.node_nodeID = n.nodeID',
        "Hostname" : ' ',
        "Scaling" : ' ',
    }
    join_on_map = {
        'Kernel Version': 'INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID',
        'OS Version': 'INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID',
        'OS Name': 'INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID',
        "Firmware Version": ' ',
        "ToolChain Name": 'INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID',
        "ToolChain Version" : 'INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID',
        "Flags": 'INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID',
        "SMT" : 'INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "Cores": 'INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "Corefreq": 'INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "DDRfreq": 'INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID',
        "SKUID": ' ',
        "Hostname" : ' ',
        "Scaling" : ' ',
    }

    # Conditions for filtering x_list
    filter_x_list_map = {
        "Kernel Version": 'not empty string', 
        'OS Version': 'not empty string',
        'OS Name': 'not empty string',
        "Firmware Version": 'not empty string',
        "ToolChain Name": 'not empty string',
        "ToolChain Version" : 'not empty string',
        "Flags": 'not empty string',
        "SMT" : 'greater than zero',
        "Cores": 'greater than zero',
        "Corefreq": 'greater than zero',
        "DDRfreq": 'greater than zero',
        "SKUID": 'not empty string',
        "Hostname": 'not empty string',
        "Scaling" : '',
    }

    # List of lists
    # Each list has entries for a single CPU Manufacturer
    x_list_list = []
    y_list_list = []
    originID_list_list = []
    legend_list = []
    x_list_order = []

    if xParameter == "Scaling":
        # [dual socket, single socket, 1/2 socket, 1/4th socket, 1/8th  socket, 2 cores, single core, single thread]
        initial_x_list = [3, 2, 7, 6, 5, 8, 1, 0]
    else:
        # Get initial_x_list by excecuting the query
        X_LIST_QUERY = "SELECT DISTINCT " + parameter_map[xParameter] + " as \'" + parameter_map[xParameter] + \
                        """\' FROM origin o INNER JOIN result r ON o.originID = r.origin_originID 
                        INNER JOIN testdescriptor t ON t.testdescriptorID=o.testdescriptor_testdescriptorID
                        INNER JOIN subtest s ON r.subtest_subtestID=s.subtestID 
                        INNER JOIN display disp ON  r.display_displayID = disp.displayID """ + \
                        join_on_map_for_x_query[xParameter] + \
                        " WHERE t.testname=\'""" + testname + "\'" + INPUT_FILTER_CONDITION + ";"

        logging.debug("Printing XLIST QUERY")
        logging.debug(X_LIST_QUERY)

        x_df = pd.read_sql(X_LIST_QUERY, db)
        logging.debug("X_Dataframe = {}".format(x_df))
        initial_x_list = sorted(x_df[parameter_map[xParameter]].to_list())
        logging.debug("X_LIST = {}".format(initial_x_list))

    if filter_x_list_map[xParameter] == 'not empty string':
        # Convert each element to type "str"
        initial_x_list = list(map(lambda x: str(x).strip(), initial_x_list))

        # Remove ALL the entries which are '' in the list 
        initial_x_list = list(filter(lambda x: x != '', initial_x_list))
    elif filter_x_list_map[xParameter] == 'greater than zero':
        initial_x_list = list(filter(lambda x: x > 0, initial_x_list))
    else:
        pass

    logging.debug("initial_x_list = {}".format(initial_x_list))

    if min_or_max_list[index] == '0':
        TIMELINE_QUERY = "SELECT r.number, o.originID, o.testdate, b.smt as 'smt_filter', n.skuidname as 'skuidname_legend', s.resultype as 'resultype_filter', r.isvalid, " + \
                        parameter_map[xParameter] + """ FROM origin o 
                        INNER JOIN result r ON o.originID = r.origin_originID 
                        INNER JOIN display disp ON  r.display_displayID = disp.displayID 
                        INNER JOIN testdescriptor t ON t.testdescriptorID = o.testdescriptor_testdescriptorID 
                        INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID 
                        INNER JOIN node n ON hw.node_nodeID = n.nodeID 
                        INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID 
                        INNER JOIN subtest s ON r.subtest_subtestID=s.subtestID """ + join_on_map[xParameter] + \
                        " WHERE t.testname = \'" + testname + "\' AND r.number > 0 " + \
                        " AND disp.qualifier LIKE \'%" + qualifier_list[index] + "%\'" + \
                        INPUT_FILTER_CONDITION + ";"
    else:
        TIMELINE_QUERY = "SELECT r.number, o.originID, o.testdate, b.smt as 'smt_filter', n.skuidname as 'skuidname_legend', s.resultype as 'resultype_filter', r.isvalid, " + \
                        parameter_map[xParameter] + """ FROM origin o 
                        INNER JOIN result r ON o.originID = r.origin_originID 
                        INNER JOIN display disp ON  r.display_displayID = disp.displayID 
                        INNER JOIN testdescriptor t ON t.testdescriptorID = o.testdescriptor_testdescriptorID 
                        INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID 
                        INNER JOIN node n ON hw.node_nodeID = n.nodeID 
                        INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID 
                        INNER JOIN subtest s ON r.subtest_subtestID=s.subtestID """ + join_on_map[xParameter] + \
                      " WHERE t.testname = \'" + testname + \
                        "\' AND disp.qualifier LIKE \'%" + qualifier_list[index] + "%\'" + \
                        INPUT_FILTER_CONDITION + ";"

    results_df = pd.read_sql(TIMELINE_QUERY, db)
    logging.debug(TIMELINE_QUERY)
    logging.debug(results_df.shape)

    if not results_df.empty:
        # Remove results which are not valid
        results_df = results_df[results_df['isvalid'] == 1].reset_index(drop=True)
        del results_df['isvalid']
        logging.debug(results_df.shape)

        # Convert to actual result type string
        results_df['resultype_filter'] = results_df['resultype_filter'].apply(lambda x : result_type_map.get(x, "Unkown"))
        if result_type_filter:
            results_df = results_df[results_df['resultype_filter'] == result_type_filter].reset_index(drop=True)

        del results_df['resultype_filter']

        # Apply SMT filter
        results_df = results_df[results_df['smt_filter'] == int(smt_filter)].reset_index(drop=True)

        # If results_df is not empty after applying reusult_type_filter
        if not results_df.empty:
            results_df['skuidname_legend'] = results_df['skuidname_legend'].apply(lambda x: x.strip())

            # Only skuidnames which are in sku_definition.ini will be shown
            valid_skuidnames = list(itertools.chain(*[sku_parser.get(section, 'SKUID').replace('\"', '').split(',') \
                                for section in sku_parser.sections()]))
            results_df = results_df[results_df['skuidname_legend']
                                .apply(lambda x: x.strip() in valid_skuidnames)].reset_index(drop=True)

            # Filter on initial_x_list
            param_name = parameter_map[xParameter]
            param_name = param_name[param_name.find('.')+1:]

            results_df = results_df[results_df[param_name]
                                .apply(lambda x: str(x).strip().upper() in [str(e).strip() for e in map(str.upper, list(map(str, initial_x_list)))])].reset_index(drop=True)

            # Convert skuidname to corresponding section in sku_definition.ini
            results_df['skuidname_legend'] = results_df['skuidname_legend'].apply(lambda x: skuid_cpu_map[x])

            # Filter on SKU
            results_df = results_df[results_df['skuidname_legend'] == sku_filter].reset_index(drop=True)
            if not results_df.empty:
                # Convert pandas.Timestamp to a format - "month_name-year"
                results_df['test_month_year'] = results_df['testdate'].apply(lambda ts: month_name_map[ts.month] + '-' + str(ts.year))

                # Get Min or Max results and group them w.r.t. skuidname_legend and x parameter
                if min_or_max_list[index] == '0':
                    idx = results_df.groupby(by=['test_month_year', 'skuidname_legend', param_name])['number'].idxmin()
                else:
                    idx = results_df.groupby(by=['test_month_year', 'skuidname_legend', param_name])['number'].idxmax()

                results_df = results_df.loc[idx].sort_values(by=['testdate', param_name, 'skuidname_legend']).reset_index(drop=True)

                # Delete testdate. No longer needed
                del results_df['testdate']

                # The order according to which X axis of the timeline graph is sorted
                x_list_order = results_df['test_month_year'].tolist()

                legend_list = [x for x in results_df[param_name].unique().tolist()]

                results_df = results_df.set_index(param_name)

                # Extract the lists from the dataframe
                # Always pass a list to .loc function to get Dataframe as the result
                x_list_list = [results_df.loc[[param]]['test_month_year'].tolist() for param in legend_list]
                y_list_list = [results_df.loc[[param]]['number'].tolist() for param in legend_list]
                originID_list_list = [results_df.loc[[param]]['originID'].tolist() for param in legend_list]

    # Get the unit for the selected yParamter (qualifier)
    UNIT_QUERY = """SELECT disp.qualifier, disp.unit FROM origin o INNER JOIN testdescriptor t 
                    ON t.testdescriptorID=o.testdescriptor_testdescriptorID  INNER JOIN result r 
                    ON o.originID = r.origin_originID  INNER JOIN display disp ON  r.display_displayID = disp.displayID 
                    where t.testname = \'""" + testname + "\' and disp.qualifier LIKE \'%" + yParameter.strip() +"%\' limit 1;"
    unit_df = pd.read_sql(UNIT_QUERY, db)

    logging.debug("Printing UNIT Query")
    logging.debug(UNIT_QUERY)
    try:
        y_axis_unit = unit_df['unit'][0]
    except Exception as error_message:
        y_axis_unit = "Unknown Unit"
        logging.debug("= {}".format(error_message))

    response = {
        'x_list_list': x_list_list, 
        'y_list_list': y_list_list,
        'y_axis_unit': y_axis_unit,
        'xParameter': xParameter,
        'yParameter': yParameter,
        'originID_list_list': originID_list_list,
        'legend_list' : legend_list,
        'x_list_order' : x_list_order,
        'graphTitle' : "Timeline Graphs - " + testname + " SMT - " + smt_filter + " SKU - " + sku_filter + " " + result_type_filter
    }

    logging.debug("Printing Final response")
    logging.debug(response)

    # close the database connection
    try:
        logging.debug("CLOSING CONNECTION FOR timeline_graph {}".format(testname))
        db.close()
    except:
        pass

    print("timeline_graph took {} seconds".format(time.time() - start_time))

    return response

def parallel_get_best_results(params, **kwargs):
    start_time = time.time()

    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER, passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    # The lists to be filled by the end of this function (and to be returned)
    x_list = []
    y_list = []
    originID_list = []

    #Unpacking of the tuple
    test_name, test_section, qualifier, higher_is_better = params

    # Other Arguments
    results_metadata_parser = kwargs['results_metadata_parser']
    FROM_DATE_FILTER = kwargs['FROM_DATE_FILTER']
    TO_DATE_FILTER = kwargs['TO_DATE_FILTER']
    result_type_filter = kwargs['result_type_filter']
    normalized_wrt = kwargs['normalized_wrt']

    sku_file_path = './config/sku_definition.ini'
    sku_parser = configparser.ConfigParser()
    sku_parser.read(sku_file_path)

    # Dictionary mapping from 'skuidname' : 'server_cpu_name'
    # Example 'Cavium ThunderX2(R) CPU CN9980 v2.2 @ 2.20 GHz' : Marvell TX2-B2
    skuid_cpu_map = OrderedDict({section: sku_parser.get(section, 'SKUID').replace('\"', '').split(',') for section in sku_parser.sections()})

    # Fill the sku_cpu_map with all "sku->section" mapping entries
    for section in sku_parser.sections():
        skus = sku_parser.get(section, 'SKUID').replace('\"','').split(',')
        for sku in skus:
            skuid_cpu_map[sku] = section

    # Get input_filter_condition by calling the function
    input_filters_list = results_metadata_parser.get(test_section, 'default_input') \
                                                .replace('\"', '').split(',')
    INPUT_FILTER_CONDITION = get_input_filter_condition(test_section, input_filters_list, \
                                wiki_description_file="./config/best_of_all_graph.ini")

    if higher_is_better == "0":
        BEST_RESULT_QUERY = """SELECT t.testname, r.number, o.originID, n.skuidname, s.resultype as 'resultype_filter', r.isvalid from origin o 
                            inner join result r on r.origin_originID = o.originID 
                            inner join testdescriptor t on t.testdescriptorID = o.testdescriptor_testdescriptorID 
                            inner join hwdetails hw on hw.hwdetailsID = o.hwdetails_hwdetailsID 
                            inner join node n on n.nodeID = hw.node_nodeID 
                            INNER JOIN display disp ON  r.display_displayID = disp.displayID 
                            INNER JOIN subtest s ON r.subtest_subtestID=s.subtestID 
                            where t.testname = \'""" + test_name + "\' AND r.number > 0 " + \
                            " AND disp.qualifier LIKE \'%" + qualifier + "%\' " + \
                            INPUT_FILTER_CONDITION + FROM_DATE_FILTER + TO_DATE_FILTER + ";"
    else:
        BEST_RESULT_QUERY = """SELECT t.testname, r.number, o.originID, n.skuidname, s.resultype as 'resultype_filter', r.isvalid from origin o 
                            inner join result r on r.origin_originID = o.originID 
                            inner join testdescriptor t on t.testdescriptorID = o.testdescriptor_testdescriptorID 
                            inner join hwdetails hw on hw.hwdetailsID = o.hwdetails_hwdetailsID 
                            inner join node n on n.nodeID = hw.node_nodeID 
                            INNER JOIN display disp ON  r.display_displayID = disp.displayID 
                            INNER JOIN subtest s ON r.subtest_subtestID=s.subtestID 
                            where t.testname = \'""" + test_name + \
                            "\' AND disp.qualifier LIKE \'%" + qualifier + "%\' " + \
                            INPUT_FILTER_CONDITION + FROM_DATE_FILTER + TO_DATE_FILTER + ";"


    results_df = pd.read_sql(BEST_RESULT_QUERY, db)
    results_df = results_df[results_df['isvalid'] == 1].reset_index(drop=True);
    del results_df['isvalid']

    if not results_df.empty:
        # Convert to actual result type string
        results_df['resultype_filter'] = results_df['resultype_filter'].apply(lambda x : result_type_map.get(x, "Unkown"))
        if result_type_filter:
            results_df = results_df[results_df['resultype_filter'] == result_type_filter].reset_index(drop=True)

        del results_df['resultype_filter']

        # If results_df is not empty after applying reusult_type_filter
        if not results_df.empty:
            # Strip all the skuidnames
            results_df['skuidname'] = results_df['skuidname'].apply(lambda x: x.strip())

            # Only valid skuidnames
            valid_skuidnames = list(itertools.chain(*[sku_parser.get(section, 'SKUID').replace('\"', '').split(',') \
                                                        for section in sku_parser.sections()]))
            results_df = results_df[results_df['skuidname'].apply(lambda x: x.strip() in valid_skuidnames)].reset_index(drop=True)

            # Convert skuidname to corresponding SKU section
            results_df['skuidname'] = results_df['skuidname'].apply(lambda x: skuid_cpu_map[x])

            if higher_is_better == "0":
                idx = results_df.groupby(by=['skuidname'])['number'].idxmin()
            else:
                idx = results_df.groupby(by=['skuidname'])['number'].idxmax()

            # Get best results for each sku
            results_df = results_df.loc[idx].sort_values(by=['skuidname']).reset_index(drop=True)

            if 'skuidname' in results_df.columns:
                if normalized_wrt in results_df['skuidname'].tolist():
                    refnum = results_df.loc[results_df.skuidname == normalized_wrt]['number'].reset_index(drop=True)[0]

                    if higher_is_better == "0":
                        results_df['number'] = results_df['number'].apply(lambda x: refnum/x)
                    else:
                        results_df['number'] = results_df['number'].apply(lambda x: x/refnum)

                    results_df = results_df.loc[results_df.skuidname != normalized_wrt]

                    # Convert test_name to test_section
                    results_df['testname'] = results_df['testname'].apply(lambda x: test_section)

                    logging.debug("Parallel best results for {} took {} seconds".format(test_name, time.time() - start_time))

                    return results_df
        
    logging.debug("Parallel best results - returning Empty DataFrame, for {} took {} seconds".format(test_name, time.time() - start_time))
    # Return empty dataframe. It becomes easier while filtering
    return pd.DataFrame()

@app.route('/best_of_all_graph', methods=['POST'])
def best_of_all_graph():
    # Just for testing the speed
    start_time = time.time()

    wiki_metadata_file_path = './config/best_of_all_graph.ini'
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(wiki_metadata_file_path)

    all_test_sections = results_metadata_parser.sections()

    sku_file_path = './config/sku_definition.ini'
    sku_parser = configparser.ConfigParser()
    sku_parser.read(sku_file_path)

    data = request.get_json()
    
    logging.debug(" = {}".format(data))

    # Apply DATE - FILTERS
    from_date = data['from_date_filter']
    to_date = data['to_date_filter']
    logging.debug(" = {}".format(from_date))
    logging.debug(" = {}".format(to_date))

    FROM_DATE_FILTER = " "
    TO_DATE_FILTER = " "
    if from_date:
        FROM_DATE_FILTER = " and o.testdate > \'" + from_date + " 00:00:00\' "

    if to_date:
        TO_DATE_FILTER = " and o.testdate < \'" + to_date + " 23:59:59\' "

    normalized_wrt = data['normalizedWRT']

    # Result type filter eg. dual socket
    result_type_filter = data['resultTypeFilter']

    # result_type_filter = "dual socket"
    logging.debug("Result type filter = '{}' {}".format(result_type_filter, type(result_type_filter)))

    # If No filters are applied, select all tests
    try:
        test_name_list = [test_name.strip() for test_name in data['test_name_list']]

        # If test_name_list is empty, read everything
        if not test_name_list:
            raise Exception
    except:
        # Read all test names from .ini file
        test_name_list = [results_metadata_parser.get(section, 'testname').strip() for section in all_test_sections]

    # A list of 'sections' corresponding to filtered(selected) tests in test_name_list
    test_sections_list = sorted([section for section in all_test_sections if results_metadata_parser.get(section,'testname') in test_name_list])

    # Modify test_name_list according to test_sections_list
    # This is a necessary step since we have multiple sections for the same 'testname'
    test_name_list = sorted([results_metadata_parser.get(section, 'testname') for section in test_sections_list])

    logging.debug("LENGTH of sections i.e. no of benchmarks = ", len(all_test_sections))
    logging.debug("Printing selected testnames list", test_name_list, len(test_name_list))
    logging.debug("\n\nPrinting corresponding sections list", test_sections_list, len(test_sections_list))

    # The list of the first qualifier for all tests
    # and the corresponding higher_is_better
    qualifier_list = [results_metadata_parser.get(section, 'fields').replace('\"', '').split(',')[0] for section in test_sections_list]
    higher_is_better_list = [results_metadata_parser.get(section, 'higher_is_better').replace('\"', '').replace(' ','').split(',')[0] for section in test_sections_list]

    # Remove all the entries where fields = ""
    while True:
        try:
            index = qualifier_list.index('')
            qualifier_list.remove(qualifier_list[index])
            higher_is_better_list.remove(higher_is_better_list[index])
            test_name_list.remove(test_name_list[index])
            test_section_list.remove(test_section_list[index])
        except:
            logging.debug("DONE REMOVING")
            break

    logging.debug(" = {}".format(qualifier_list))
    logging.debug(" = {}".format(higher_is_better_list))

    # Get colour and skuid_list for reference Cpu 
    reference_color = sku_parser.get(normalized_wrt, 'color').replace('\"', '').split(',')[0]
    reference_skuid_list = sku_parser.get(normalized_wrt, 'SKUID').replace('\"', '').split(',')

    ######################################################################################################################

    x_list_list = []
    y_list_list = []
    originID_list_list = []
    server_cpu_list = []
    color_list = []

    # Excecute parallel_get_reference_results parallely with multiprocessing "pool"
    pool = multiprocessing.Pool(processes=num_processes)

    start_time2 = time.time()
    try:
        best_results_data = pool.map(partial(parallel_get_best_results, results_metadata_parser=results_metadata_parser, \
                                FROM_DATE_FILTER = FROM_DATE_FILTER, TO_DATE_FILTER = TO_DATE_FILTER, \
                                result_type_filter=result_type_filter, normalized_wrt=normalized_wrt), \
                                zip(test_name_list, test_sections_list, qualifier_list, higher_is_better_list)) 
    except:
        pass
    finally:
        logging.debug("Closing pool")
        pool.close()
        pool.join()

    # Remove all the empty dataframes from the list
    best_results_data = [x for x in best_results_data if not x.empty]

    # Concatenate all the dfs into a single df
    if best_results_data:
        best_results_df = pd.concat(best_results_data).sort_values(by=['skuidname', 'testname']).reset_index(drop=True)
    else:
        best_results_df = pd.DataFrame()

    # If best_results_df is not empty
    if not best_results_df.empty:

        print("Sorting best results df acc to skuidname")
        # Sort by skuidname according to the sku_definition.ini file
        sku_categories = sku_parser.sections()
        ordered_skus = pd.Categorical(best_results_df['skuidname'].tolist(), categories = sku_categories, ordered=True)
        best_results_df['skuidname'] = pd.Series(ordered_skus)
        best_results_df = best_results_df.sort_values(by='skuidname')

        # Get unique skuidname entries
        server_cpu_list = [x for x in best_results_df['skuidname'].unique().tolist()]

        best_results_df = best_results_df.set_index('skuidname')

        # Extract the lists from the dataframe
        # Always pass a list to .loc function to get Dataframe as the result
        x_list_list = [best_results_df.loc[[skuidname]]['testname'].tolist() for skuidname in server_cpu_list]
        y_list_list = [best_results_df.loc[[skuidname]]['number'].tolist() for skuidname in server_cpu_list]
        originID_list_list = [best_results_df.loc[[skuidname]]['originID'].tolist() for skuidname in server_cpu_list]

        for section in server_cpu_list:
            color_list.extend(sku_parser.get(section, 'color').replace('\"','').split(','))

    ######################################################################################################################

    response = {
        'server_cpu_list': server_cpu_list,
        'color_list': color_list,
        'x_list_list': x_list_list, 
        'y_list_list': y_list_list,
        'originID_list_list': originID_list_list,
        'y_axis_unit': "ratio",
        'xParameter': "",
        'yParameter': "",
        'reference_color' : reference_color,
        'normalized_wrt' : normalized_wrt,
    }

    print("Best of All Graph took {} seconds".format(time.time() - start_time))

    return response

def parallel_compute_heatmap_zll(param, **kwargs):
    graph_name = kwargs['graph_name']

    if graph_name == 'cpu_heatmap':
        df = param
        return df['%busy'].tolist()   
    elif graph_name == 'network_heatmap':
        df = param
        return df['NW_UTIL'].tolist()
    elif graph_name == 'softirq_heatmap':
        df = param
        return df['%soft'].tolist()
    elif graph_name == 'ram_heatmap':
        node = param
        df = kwargs['ramstat_df']
        return df[node].tolist()

def parallel_compute_freq_dump_yll(param, **kwargs):
    df = kwargs['df']

    col = param

    try:
        return (list(range(df.shape[0])), df[col].tolist(), col)
    except:
        return None

def parallel_compute_iostat_yll(device_type, **kwargs):
    df = kwargs['df']
    col = kwargs['col']

    try:
        return (list(range(df.loc[device_type, col].shape[0])), df.loc[device_type, col].tolist(), device_type)
    except:
        return None

# API Endpoint for CPU Utilization graphs
@app.route('/cpu_utilization_graphs', methods=['POST'])
def cpu_utilization_graphs():
    logging.debug("Got request for heatmap")
    start_time = time.time()

    data = request.get_json()
    logging.debug("DATA = {}".format(data))
    numCPUs = data['numCPUs']

    # Get data for heatmap
    heatmap_data = {
        'graph_type' : 'heatmap',
        'x_list' : [],
        'y_list' : [],
        'z_list_list' : [],
        'xParameter' : 'Timestamp',
        'yParameter' : 'Cores'
    }

    softirq_heatmap_data = {
        'graph_type' : 'heatmap',
        'x_list' : [],
        'y_list' : [],
        'z_list_list' : [],
        'xParameter' : 'Timestamp',
        'yParameter' : 'Cores'          
    }

    # Get data for line graph from all_cores_df
    line_graph_data = {
        'graph_type' : 'line',
        'x_list_list' : [],
        'y_list_list' : [],     #list_list because the JS function is written for multiple lines
        'legend_list' : [],
        'xParameter' : 'Timestamp',
        'yParameter' : '% Utilization'
    }

    # Get data for stack graph from average_cpu_ut_df
    stack_graph_data = {
        'graph_type' : 'stack',
        'x_list' : [],
        'y_list_list' : [],
        'legend_list' : [],
        'xParameter' : 'Cores',
        'yParameter' : 'AVG. % Utilization',
    }

    # Generate nas_path from received data
    nas_path = "/mnt/nas/dbresults/" + data['jobname'] + '/' + data['runID'] + '/results/' + numCPUs
    cpu_file = nas_path + '/CPU_heatmap.csv'
    logging.debug("NAS PATH = {}".format(cpu_file))

    start_time2 = time.time()
    # Check if CPU_heatmap.csv exists
    if os.path.isfile(cpu_file):
        cpu_utilization_df = pd.read_csv(cpu_file, usecols=['timestamp','CPU', '%idle', '%soft', '%usr', '%nice', '%sys', '%iowait', '%irq', '%steal', '%guest', '%gnice'])

        logging.debug("Reading CPU and N/W CSV files took {} seconds".format(time.time() - start_time2))

        start_time3 = time.time()
        try:
            # Get data for stack graph from average_cpu_ut_df

            cpu_utilization_df = cpu_utilization_df.set_index('timestamp')
            # For stack graph having average entries of all Cores and average of 'all'
            average_cpu_ut_df = cpu_utilization_df.loc['Average:']
            # Drop all those columns
            cpu_utilization_df = cpu_utilization_df.drop('Average:')

            # Columns list for stack graph
            stack_cols_list = ['%idle', '%soft', '%usr', '%nice', '%sys', '%iowait', '%irq', '%steal', '%guest', '%gnice']

            stack_graph_data['x_list'] = average_cpu_ut_df['CPU'].tolist()

            for col in stack_cols_list:
                stack_graph_data['y_list_list'].append(average_cpu_ut_df[col].tolist())
                stack_graph_data['legend_list'].append(col)
            # Stack graph is done
        except:
            pass
        finally:
            # Reset index
            cpu_utilization_df = cpu_utilization_df.reset_index()

        logging.debug("Stack graph took {} seconds".format(time.time() - start_time3))

        start_time4 = time.time()
        # Calculate %busy by 100-%idle
        cpu_utilization_df['%busy'] = cpu_utilization_df['%idle'].apply(lambda x: 100 - x)
        cpu_utilization_df.pop('%idle')

        logging.debug("Busy Column generation took {} seconds".format(time.time() - start_time4))

        start_time5 = time.time()
        # Set 'CPU' as index
        cpu_utilization_df = cpu_utilization_df.set_index('CPU')

        try:
            # AVG. Data of all cores at all timestamps
            all_cores_df = cpu_utilization_df.loc['all']
            # Drop all those columns
            cpu_utilization_df = cpu_utilization_df.drop('all')
        except:
            all_cores_df = pd.DataFrame()

        # Reset index
        cpu_utilization_df = cpu_utilization_df.reset_index()

        logging.debug("Deleting 'all' cores columns took {} seconds".format(time.time() - start_time5))

        logging.debug("PRINTING CPU Utilization DF")
        logging.debug(cpu_utilization_df.columns)

        start_time6 = time.time()
        heatmap_data['x_list'] = cpu_utilization_df['timestamp'].unique().tolist()
        heatmap_data['y_list'] = cpu_utilization_df['CPU'].unique().tolist()

        start_time7 = time.time()
        # list of lists. Length = unique timestamps = len(heatmap_data['x_list'])
        # Each list has data for one timestamp
        # Length of each list = No. of Cores
        # x in range cpu_util_df['%busy'] with jumps of length = no of cores
        pool = multiprocessing.Pool(num_processes)

        try:
            heatmap_data['z_list_list'] = pool.map(partial(parallel_compute_heatmap_zll, graph_name='cpu_heatmap'), \
                        np.array_split(cpu_utilization_df, (cpu_utilization_df.shape[0]/len(heatmap_data['y_list']))))
        except:
            pass
        finally:
            logging.debug("Closing pool")
            pool.close()
            pool.join()

        logging.debug("Z LIST LIST took {} seconds".format(time.time() - start_time7))

        start_time8 = time.time()
        # Take transpose of the list_list
        # Because plotly.js plots it left->right, then top->bottom
        # So now
        # list of lists. Length = No. of Cores
        # Each list has data for one Core
        # Length of each list = unique timestamps = len(heatmap_data['x_list'])
        heatmap_data['z_list_list'] = np.array(heatmap_data['z_list_list']).T.tolist()

        logging.debug("TRANSPOSE of Z list list took {} seconds".format(time.time() - start_time8))

        logging.debug("CPU UTIL HEATMAP overall took {} seconds".format(time.time() - start_time6))
        # CPU Util Heatmap is done

        start_time12 = time.time()
        softirq_heatmap_data['x_list'] = cpu_utilization_df['timestamp'].unique().tolist()
        softirq_heatmap_data['y_list'] = cpu_utilization_df['CPU'].unique().tolist()

        start_time13 = time.time()


        pool = multiprocessing.Pool(num_processes)
        try:
            softirq_heatmap_data['z_list_list'] = pool.map(partial(parallel_compute_heatmap_zll, graph_name='softirq_heatmap'), \
                        np.array_split(cpu_utilization_df, (cpu_utilization_df.shape[0]/len(softirq_heatmap_data['y_list']))))
        except:
            pass
        finally:
            logging.debug("Closing pool")
            pool.close()
            pool.join()


        logging.debug("Z LIST LIST took {} seconds".format(time.time() - start_time13))
        # Take transpose of the list_list 
        start_time14 = time.time()
        softirq_heatmap_data['z_list_list'] = np.array(softirq_heatmap_data['z_list_list']).T.tolist()
        logging.debug("TRANSPOSE Z list list took {} seconds".format(time.time() - start_time14))

        logging.debug("SoftIRQ Heatmap overall took {} seconds".format(time.time() - start_time12))
        #SoftIRQ heatmap is done

        start_time15 = time.time()
        # Get data for line graph from all_cores_df
        if not all_cores_df.empty:
            line_graph_data['x_list_list'].append(all_cores_df['timestamp'].tolist())
            line_graph_data['y_list_list'].append(all_cores_df['%busy'].tolist())
            line_graph_data['legend_list'].append('Avg CPU Utilization')


            line_graph_data['x_list_list'].append(all_cores_df['timestamp'].tolist())
            line_graph_data['y_list_list'].append(all_cores_df['%soft'].tolist())
            line_graph_data['legend_list'].append('%soft')
            #%idle', '%soft', '%usr', '%nice', '%sys', '%iowait', '%irq', '%steal', '%guest', '%gnice'

            line_graph_data['x_list_list'].append(all_cores_df['timestamp'].tolist())
            line_graph_data['y_list_list'].append(all_cores_df['%usr'].tolist())
            line_graph_data['legend_list'].append('%usr')

            line_graph_data['x_list_list'].append(all_cores_df['timestamp'].tolist())
            line_graph_data['y_list_list'].append(all_cores_df['%nice'].tolist())
            line_graph_data['legend_list'].append('%nice')

            line_graph_data['x_list_list'].append(all_cores_df['timestamp'].tolist())
            line_graph_data['y_list_list'].append(all_cores_df['%sys'].tolist())
            line_graph_data['legend_list'].append('%sys')

            line_graph_data['x_list_list'].append(all_cores_df['timestamp'].tolist())
            line_graph_data['y_list_list'].append(all_cores_df['%iowait'].tolist())
            line_graph_data['legend_list'].append('%iowait')

            line_graph_data['x_list_list'].append(all_cores_df['timestamp'].tolist())
            line_graph_data['y_list_list'].append(all_cores_df['%irq'].tolist())
            line_graph_data['legend_list'].append('%irq')

            line_graph_data['x_list_list'].append(all_cores_df['timestamp'].tolist())
            line_graph_data['y_list_list'].append(all_cores_df['%steal'].tolist())
            line_graph_data['legend_list'].append('%steal')

            line_graph_data['x_list_list'].append(all_cores_df['timestamp'].tolist())
            line_graph_data['y_list_list'].append(all_cores_df['%guest'].tolist())
            line_graph_data['legend_list'].append('%guest')

            line_graph_data['x_list_list'].append(all_cores_df['timestamp'].tolist())
            line_graph_data['y_list_list'].append(all_cores_df['%gnice'].tolist())
            line_graph_data['legend_list'].append('%gnice')
    else:
        print("File CPU_heatmap.csv does not exist. Skipping CPU Util graphs")

    # Network util graphs
    network_heatmap_data = {
        'graph_type' : 'heatmap',
        'x_list' : [],
        'y_list' : [],
        'z_list_list' : [],
        'xParameter' : 'Timestamp',
        'yParameter' : 'Interface'
    }

    network_file = nas_path + '/ethperc.csv'
    # Check if ethperc.csv exists
    if os.path.isfile(network_file):
        network_utilization_df = pd.read_csv(network_file,usecols=['Time','Interface','NW_UTIL'])

        start_time9 = time.time()
        network_heatmap_data['x_list'] = network_utilization_df['Time'].unique().tolist()

        network_heatmap_data['y_list'] = network_utilization_df['Interface'].unique().tolist()

        start_time10 = time.time()

        pool = multiprocessing.Pool(num_processes)
        try:
            network_heatmap_data['z_list_list'] = pool.map(partial(parallel_compute_heatmap_zll, graph_name='network_heatmap'), \
                                                    np.array_split(network_utilization_df, (network_utilization_df.shape[0]/len(network_heatmap_data['y_list']))))
        except:
            pass
        finally:
            logging.debug("Closing pool")
            pool.close()
            pool.join()

        logging.debug("Z LIST LIST took {} seconds".format(time.time() - start_time10))

        start_time11 = time.time()
        network_heatmap_data['z_list_list'] = np.array(network_heatmap_data['z_list_list']).T.tolist()
        logging.debug("TRANSPOSE Z list list took {} seconds".format(time.time() - start_time11))

        logging.debug("Network Heatmap overall took {} seconds".format(time.time() - start_time9))
        #Network heatmap is done


        #line_graph_data['x_list_list'].append(network_utilization_df['Time'].unique().tolist())
        for intface in network_utilization_df['Interface'].unique().tolist():
            #logging.debug(network_utilization_df.query('Interface'==intface)['NW_UTIL'].tolist())
            line_graph_data['x_list_list'].append(network_utilization_df['Time'].unique().tolist())
            line_graph_data['y_list_list'].append(network_utilization_df[network_utilization_df['Interface']==intface]['NW_UTIL'].tolist())
            nw_util_str = intface
            line_graph_data['legend_list'].append(nw_util_str)
            #logging.debug(network_utilization_df[network_utilization_df['Interface']==intface]['NW_UTIL'].tolist())
        logging.debug("Line Graph overall took {} seconds".format(time.time() - start_time15))
        # Line graph data is done
        # All network util graphs are done
    else:
        print("File ethperc.csv does not exist. Skipping Network graphs")


    # Do NOT change key names.
    # Changing them will require changes in HTML code
    cpu_ut_graphs_data = OrderedDict()
    
    cpu_ut_graphs_data['CPU %busy heatmap'] = heatmap_data
    cpu_ut_graphs_data['CPU %softirq heatmap'] = softirq_heatmap_data
    cpu_ut_graphs_data['network_heatmap_data'] = network_heatmap_data
    cpu_ut_graphs_data['%CPU Utilization Multi-line Graph'] = line_graph_data
    cpu_ut_graphs_data['%CPU Utilization Stack graph'] = stack_graph_data
    
    # Freq dump graphs
    freq_dump_file = nas_path + '/freq_dump.csv'
    # Check if freq_dump.csv exists
    if os.path.isfile(freq_dump_file):
        start_time16 = time.time()

        freq_dump_df = pd.read_csv(freq_dump_file)

        # Drop all columns with any 'NaN' value
        freq_dump_df.dropna(axis=1, how='any', inplace=True)

        memnet_freq_line_graph_data = {
            'graph_type' : 'line',
            'x_list_list' : [],
            'y_list_list' : [],     #list_list because the JS function is written for multiple lines
            'legend_list' : [],
            'xParameter' : 'Timestamp',
            'yParameter' : 'Core and memnet frequency'
        }

        voltage_line_graph_data = {
            'graph_type' : 'line',
            'x_list_list' : [],
            'y_list_list' : [],     #list_list because the JS function is written for multiple lines
            'legend_list' : [],
            'xParameter' : 'Timestamp',
            'yParameter' : 'Voltage in V'
        }

        power_stack_graph_data = {
            'graph_type' : 'stack',
            'x_list' : [],
            'y_list_list' : [],
            'legend_list' : [],
            'xParameter' : '',
            'yParameter' : 'Power in W'
        }

        temperature_line_graph_data = {
            'graph_type' : 'line',
            'x_list_list' : [],
            'y_list_list' : [],     #list_list because the JS function is written for multiple lines
            'legend_list' : [],
            'xParameter' : '',
            'yParameter' : 'Temperature in °C'
        }


        no_of_nodes = freq_dump_df['Node'].unique().tolist()

        all_voltage_columns = ['core-voltage', 'mem-voltage']
        all_power_columns = ['core-power','mem-power','sram-power','soc-power']
        all_temperature_columns = ['temperature']
        
        memnet_columns = [x for x in freq_dump_df.columns.tolist() \
            if x not in all_power_columns and x not in all_temperature_columns \
            and x not in all_voltage_columns and x != "Node"]

        voltage_columns = [x for x in all_voltage_columns \
                            if x in freq_dump_df.columns.tolist()]
        power_columns = [x for x in all_power_columns \
                            if x in freq_dump_df.columns.tolist()]
        temperature_columns = [x for x in all_temperature_columns \
                                if x in freq_dump_df.columns.tolist()]

        freq_dump_df = freq_dump_df.set_index('Node')

        logging.debug("NO OF NODES = ", no_of_nodes, type(no_of_nodes[0]))

        for node in no_of_nodes:
            pool = multiprocessing.Pool(num_processes)
            logging.debug("PRINTING NODE = ", node)
            df = freq_dump_df.loc[int(node)]
            
            try:
                # Memnet freq
                temp_data = pool.map(partial(parallel_compute_freq_dump_yll, df=df, graph_name='memnet_graph'), memnet_columns)
                # Filter out NoneType elements
                temp_data = list(filter(None, temp_data))

                memnet_freq_line_graph_data['x_list_list'].extend([x[0] for x in temp_data])
                memnet_freq_line_graph_data['y_list_list'].extend([x[1] for x in temp_data])
                # Legend list is the list of columns
                memnet_freq_line_graph_data['legend_list'].extend(['Node-'+str(node)+'-'+x[2] for x in temp_data])

                # Power
                temp_data = pool.map(partial(parallel_compute_freq_dump_yll, df=df, graph_name='power_graph'), power_columns)
                # Filter out NoneType elements
                temp_data = list(filter(None, temp_data))

                power_stack_graph_data['x_list'] = list(range(df.shape[0]))
                power_stack_graph_data['y_list_list'].extend([x[1] for x in temp_data])
                # Legend list is the list of columns
                power_stack_graph_data['legend_list'].extend(['Node-'+str(node)+'-'+x[2] for x in temp_data])
                
                # Voltage

                temp_data = pool.map(partial(parallel_compute_freq_dump_yll, df=df, graph_name='voltage_graph'), voltage_columns)
                # Filter out NoneType elements
                temp_data = list(filter(None, temp_data))

                voltage_line_graph_data['x_list_list'].extend([x[0] for x in temp_data])
                voltage_line_graph_data['y_list_list'].extend([x[1] for x in temp_data])
                voltage_line_graph_data['legend_list'].extend(['Node-'+str(node)+'-'+x[2] for x in temp_data])

                # Temperature
                temp_data = pool.map(partial(parallel_compute_freq_dump_yll, df=df, graph_name='temperature_graph'), temperature_columns)
                # Filter out NoneType elements
                temp_data = list(filter(None, temp_data))

                temperature_line_graph_data['x_list_list'].extend([x[0] for x in temp_data])
                temperature_line_graph_data['y_list_list'].extend([x[1] for x in temp_data])
                # Legend list is the list of columns
                temperature_line_graph_data['legend_list'].extend(['Node-'+str(node)+'-'+x[2] for x in temp_data])
            except:
                pass
            finally:
                logging.debug("Closing pool")
                pool.close()
                pool.join()

        power_voltage_graph_data = {
                'graph_type' : 'combo',
                'graph_1_data' : power_stack_graph_data,
                'graph_2_data' : voltage_line_graph_data,
        }

        freq_dump_df = freq_dump_df.reset_index()
        # Add freq_dump data in context dict
        cpu_ut_graphs_data['Core frequency'] = memnet_freq_line_graph_data
        cpu_ut_graphs_data['Power & Voltage Consumption vs Timestamp'] = power_voltage_graph_data
        cpu_ut_graphs_data['Temperature'] = temperature_line_graph_data

        logging.debug("Freq dump Graphs overall took {} seconds".format(time.time() - start_time16))
    else:
        logging.debug("File freq_dump.csv does not exist. Skipping")

    # Freq dump graphs DONE

    # Ram Utilization Graphs
    ram_file = nas_path + '/ramstat.csv'
    # Check if ramstat.csv file exists
    if os.path.isfile(ram_file):
        start_time17 = time.time()
        ramstat_df = pd.read_csv(ram_file)

        ram_heatmap_data = {
            'graph_type' : 'heatmap',
            'x_list' : [],
            'y_list' : [],
            'z_list_list' : [],
            'xParameter' : 'Timestamp',
            'yParameter' : 'RAM Nodes'
        }

        # Get relevant data from dataframe
        ram_heatmap_data['x_list'] = ramstat_df['Timestamp'].unique().tolist()
        ram_heatmap_data['y_list'] = ramstat_df.columns.tolist()
        ram_heatmap_data['y_list'].remove('Timestamp')

        # Fill z_list_list for each y_list element (Each Node)
        pool = multiprocessing.Pool(num_processes)

        try:
            ram_heatmap_data['z_list_list'] = \
                        pool.map(partial(parallel_compute_heatmap_zll, \
                        graph_name='ram_heatmap', ramstat_df=ramstat_df), \
                        ram_heatmap_data['y_list'])
        except:
            pass
        finally:
            logging.debug("Closing pool")
            pool.close()
            pool.join()
        # Ram util heatmap done

        # Idea - RAM Line graph data can be derived from ram_heatmap_data. No need to recompute
        # x_list_list (line) = list of x_list(heatmap) repeated len(y_list(heatmap)) times
        # y_list_list (line) = z_list_list (heatmap)
        # legend_list (line) = y_list (heatmap)
        ram_line_graph_data = {
            'graph_type' : 'line',
            'x_list_list' : [],
            'y_list_list' : [],     #list_list because the JS function is written for multiple lines
            'legend_list' : [],
            'xParameter' : 'Timestamp',
            'yParameter' : '% Utilization'
        }

        for i in range(0, len(ram_heatmap_data['y_list'])):
            ram_line_graph_data['x_list_list'].append(ram_heatmap_data['x_list'])

        ram_line_graph_data['y_list_list'] = ram_heatmap_data['z_list_list']
        ram_line_graph_data['legend_list'] = ram_heatmap_data['y_list']

        # Add ram data in context dict
        cpu_ut_graphs_data['%RAM Utilization Heatmap'] = ram_heatmap_data
        cpu_ut_graphs_data['%RAM Utilization Line Graph'] = ram_line_graph_data
            
        logging.debug("Ram Graphs overall took {} seconds".format(time.time() - start_time17))
    else:
        logging.debug("File ramstat.csv does not exist. Skipping")

    # RAM Graphs done

    # Iostat Graphs
    iostat_file = nas_path + '/iostat.csv'
    # Check if iostat.csv file exists
    if os.path.isfile(iostat_file):
        logging.debug("IOSTAT file exists")
        start_time18 = time.time()
        iostat_df = pd.read_csv(iostat_file)

        iostat_line_graph_data = {
            'graph_type' : 'line',
            'x_list_list' : [],
            'y_list_list' : [],     #list_list because the JS function is written for multiple lines
            'legend_list' : [],
            'xParameter' : 'Timestamp',
            'yParameter' : 'kB/s'
        }

        # Get device types list
        device_types = iostat_df['Device'].unique().tolist()

        # Set 'device' as index
        iostat_df = iostat_df.set_index('Device')

        # These columns are to be selected
        column_list = ['kB_read/s', 'kB_wrtn/s']
        for col in column_list:
            pool = multiprocessing.Pool(num_processes)
            try:
                temp_data = pool.map(partial(parallel_compute_iostat_yll, df=iostat_df, col=col), device_types)
                # Filter out NoneType elements
                temp_data = list(filter(None, temp_data))

                logging.debug("Printing temp data")
                logging.debug(temp_data)
                logging.debug(type(temp_data))

                iostat_line_graph_data['x_list_list'].extend([x[0] for x in temp_data])
                iostat_line_graph_data['y_list_list'].extend([x[1] for x in temp_data])
                # Legend list is the list of columns
                iostat_line_graph_data['legend_list'].extend([col+'-'+x[2] for x in temp_data])
            except:
                pass
            finally:
                pool.close()
                pool.join()

        # Add iostat data in context dict
        cpu_ut_graphs_data['IOSTAT Line graph'] = iostat_line_graph_data

        logging.debug("IOSTAT Graphs overall took {} seconds".format(time.time() - start_time17))

    else:
        logging.debug("File iostat.csv does not exist. Skipping")

    # iostat graphs DONE

    print("Time taken for CPU utilization graphs {}".format(time.time() - start_time))

    return json.dumps(cpu_ut_graphs_data)

# One function for downloading everything as CSV
@app.route('/download_as_csv', methods=['POST'])
def download_as_csv():
    logging.debug("\n\n\n#REQUEST#########")
    logging.debug("= {}".format(request))
    logging.debug("= {}".format(request.form))
    json_data = json.loads(request.form.get('data'))

    logging.debug(" = {}".format(json_data))
    data = json_data['data']

    logging.debug(" = {}".format(data))
    logging.debug(" = {}".format(type(data)))
    logging.debug("\n\n\n")

    csv_df = pd.DataFrame(columns=data.keys());

    for column in data:
        csv_df[column] = data[column]

    # Clear the temp_download_files directory
    base_path = os.getcwd() + '/temp_download_files/'
    try:
        shutil.rmtree(base_path)    #this removes the directory too
    except:
        pass
    os.mkdir(base_path)         #So create it again

    # Save the current csv file there
    logging.debug("= {}".format(csv_df))
    logging.debug(" = {}".format(os.getcwd()))

    # "CLEAN" the filename.
    # Example -> if the file is 'Mop/s vs OSName', then replace '/' with '-per-'
    filename = base_path + json_data['filename'].replace('/', '_per_') + '.csv'
    
    logging.debug("PRINTING FILENAME AFTER = {}".format(filename))

    csv_df.to_csv(filename, index=False, header=True)
    logging.debug("\n\n\n")

    try:
        return send_file(filename,
            mimetype='text/csv',
            attachment_filename= json_data['filename'].replace('/', '_per_') + '.csv',
            as_attachment=True)

    except Exception as error_message:
        return error_message, 404

@app.route('/reports_page')
def reports_page():

    # For 'Go To Benchmark' Dropdown
    all_tests_data = get_all_tests_data(wiki_description_file='./config/best_of_all_graph.ini')

    param_list = [
        {
            'name' : 'SKUID', 'data_type' : 'string', 'input_type' : 'dropdown multiple select', 'dropdown_label' : '', 'display_by_default' : 'Yes',
            'dropdown_values_list' : [""] + all_tests_data['reference_list']
        },
        {
            'name' : 'Test Date', 'data_type' : 'date', 'input_type' : 'dropdown', 'dropdown_label' : 'Year', 'display_by_default' : 'Yes',
            'dropdown_values_list' : ["", "2020", "2019", "2018"]
        
        },
        {
            'name' : 'Scaling', 'data_type' : 'string', 'input_type' : 'dropdown', 'dropdown_label' : 'Scaling Type', 'display_by_default' : 'Yes',
            'dropdown_values_list' : [
                "", "Single Thread", "Single Socket", "Dual Socket", 
                "1/2 Socket", "1/4 Socket", "1/8 Socket", "Single core", "2 Cores",
            ]
        },
        {'name' : 'Hostname', 'data_type' : 'string', 'input_type' : 'text', 'display_by_default' : 'No'},
        {
            'name' : 'Kernel Version', 'data_type' : 'string', 'input_type' : 'text', 'display_by_default' : 'Yes',
            'criteria_op_list' : ['greater than', 'equals', 'less than']
        },
        {
            'name' : 'OS Version', 'data_type' : 'string', 'input_type' : 'text', 'display_by_default' : 'Yes',
            'criteria_op_list' : ['greater than', 'equals', 'less than']
        },
        {'name' : 'OS Name', 'data_type' : 'string', 'input_type' : 'text', 'display_by_default' : 'Yes'},
        {'name' : 'Firmware Version', 'data_type' : 'string', 'input_type' : 'text', 'display_by_default' : 'No'},
        {'name' : 'ToolChain Name', 'data_type' : 'string', 'input_type' : 'text', 'display_by_default' : 'No'},
        {'name' : 'ToolChain Version', 'data_type' : 'string', 'input_type' : 'text', 'display_by_default' : 'Yes'},
        {'name' : 'SMT', 'data_type' : 'numeric', 'input_type' : 'text', 'display_by_default' : 'No'},
        {'name' : 'Cores', 'data_type' : 'numeric', 'input_type' : 'text', 'display_by_default' : 'No'},
        {'name' : 'DDRfreq', 'data_type' : 'numeric', 'input_type' : 'text', 'display_by_default' : 'No'},
        {'name' : 'Notes', 'data_type' : 'hidden', 'input_type' : 'hidden', 'display_by_default' : 'Yes'},
    ]

    context = {'param_list':param_list}
    error = None

    return render_template('reports.html', error=error, context=context, all_tests_data=all_tests_data)

# Query the database paralelly for each 'testname'
def parallel_test_report(params, **kwargs):
    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    # Unpack the params here
    test_section, testname, INPUT_FILTER_CONDITION = params

    logging.debug("################################################################################")
    logging.debug("Processing Paralelly for {}".format(testname))
    logging.debug("Input filter condition = {}".format(INPUT_FILTER_CONDITION))

    results_metadata_file_path = './config/wiki_description.ini'
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(results_metadata_file_path)

    description_string = results_metadata_parser[testname]['description'].replace('\"', '')

    SELECT_PARAMS = kwargs['SELECT_PARAMS']
    FINAL_CRITERIA = kwargs['FINAL_CRITERIA']
    kernel_criteria = kwargs['kernel_criteria']
    os_version_criteria = kwargs['os_version_criteria']
    kernel_criteria_op = kwargs['kernel_criteria_op']
    os_version_criteria_op = kwargs['os_version_criteria_op']
    skuid_cpu_map = kwargs['skuid_cpu_map']
    all_skuidnames_criteria = kwargs['all_skuidnames_criteria']
    skuid_criteria_op = kwargs['skuid_criteria_op']
    best_results_condition = kwargs['best_results_condition']

    # An empty dataframe
    results_dataframe = pd.DataFrame()

    if best_results_condition:

        # Get 'higher_is_better' value for the first "field" of testname
        qualifier = results_metadata_parser.get(testname, 'fields').replace('\"', '').split(',')[0]
        higher_is_better = results_metadata_parser.get(testname, 'higher_is_better').replace('\"', '').replace(' ','').split(',')[0]

        # For each skuid_list
        for skuid_list in all_skuidnames_criteria:
            
            # Make SKUID_CRITERIA string
            if skuid_criteria_op == "matches":
                SKUID_CRITERIA = " AND n.skuidname IN " + str(skuid_list).replace('[','(').replace(']',')')
            else:
                SKUID_CRITERIA = " AND n.skuidname NOT IN " + str(skuid_list).replace('[','(').replace(']',')')

            # r.number and order by conditions according to higher_is_better
            if higher_is_better == '0':
                R_NUMBER = " MIN(r.number) as number, "
                ORDER_BY_CONDITION = " order by r.number "
            else:
                R_NUMBER = " MAX(r.number) as number, "
                ORDER_BY_CONDITION = " order by r.number DESC "

            if best_results_condition == 'best-results':
                LIMIT_CONDITION = " limit 1 "
            elif best_results_condition == 'top-5':
                LIMIT_CONDITION = " limit 5 "

            QUALIFIER_CONDITION = "  AND disp.qualifier LIKE \'%" + qualifier + '%\''

            GROUP_BY_CONDITION = " group by t.testname, o.originID, " + SELECT_PARAMS + \
                                     " s.description, r.number, disp.unit, disp.qualifier "

            start_time = time.time()

            RESULTS_QUERY = "SELECT t.testname, o.originID, " + SELECT_PARAMS + \
                            "s.description," + R_NUMBER + """ disp.unit, disp.qualifier 
                            FROM result r INNER JOIN subtest s ON s.subtestID=r.subtest_subtestID 
                            INNER JOIN display disp ON disp.displayID=r.display_displayID 
                            INNER JOIN origin o ON o.originID=r.origin_originID 
                            INNER JOIN testdescriptor t ON t.testdescriptorID = o.testdescriptor_testdescriptorID 
                            INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID
                            INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID
                            INNER JOIN node n ON hw.node_nodeID = n.nodeID 
                            INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID
                            INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID
                            WHERE t.testname = \'""" + testname + \
                            "\'" + INPUT_FILTER_CONDITION + FINAL_CRITERIA + SKUID_CRITERIA + " AND r.isvalid = 1 " + \
                            QUALIFIER_CONDITION + GROUP_BY_CONDITION + ORDER_BY_CONDITION + LIMIT_CONDITION + ";"

            logging.debug("\nFINAL QUERY = {}".format(RESULTS_QUERY))

            temp_df = pd.read_sql(RESULTS_QUERY, db)

            query_excecution_time = time.time() - start_time
            logging.debug("Query excecution took {} seconds".format(query_excecution_time))

            # Append the dataframe below the current dataframe
            results_dataframe = results_dataframe.append(temp_df, sort=False)

        # Replace testname with 'test_section' name
        results_dataframe = results_dataframe.reset_index(drop=True)
    else:
        RESULTS_QUERY = "SELECT t.testname, o.originID, " + SELECT_PARAMS + \
                        """s.description, r.number, disp.unit, disp.qualifier 
                        FROM result r INNER JOIN subtest s ON s.subtestID=r.subtest_subtestID 
                        INNER JOIN display disp ON disp.displayID=r.display_displayID 
                        INNER JOIN origin o ON o.originID=r.origin_originID 
                        INNER JOIN testdescriptor t ON t.testdescriptorID = o.testdescriptor_testdescriptorID 
                        INNER JOIN hwdetails hw ON o.hwdetails_hwdetailsID = hw.hwdetailsID
                        INNER JOIN bootenv b ON hw.bootenv_bootenvID = b.bootenvID
                        INNER JOIN node n ON hw.node_nodeID = n.nodeID 
                        INNER JOIN ostunings os ON o.ostunings_ostuningsID = os.ostuningsID
                        INNER JOIN toolchain tc ON o.toolchain_toolchainID = tc.toolchainID
                        WHERE t.testname = \'""" + testname + \
                        "\'" + INPUT_FILTER_CONDITION + FINAL_CRITERIA + " AND r.isvalid = 1;"

        logging.debug("\nFINAL QUERY = {}".format(RESULTS_QUERY))

        results_dataframe = pd.read_sql(RESULTS_QUERY, db)

    if 'testname' in results_dataframe:
        results_dataframe.insert(0, "Test name", [test_section for item in results_dataframe['testname']])
        del results_dataframe['testname']


    if 'testdate' in results_dataframe.columns:
        results_dataframe.insert(2, 'Test Time', [d.time() for d in results_dataframe['testdate']])
        results_dataframe.insert(2, 'Test Date', [d.date() for d in results_dataframe['testdate']])
        results_dataframe.drop('testdate', axis=1, inplace=True)

    # Filter rows on kernel_criteria
    if kernel_criteria != '':
        if kernel_criteria_op == "greater than":
            logging.debug("greater than selected for ", testname, " Kernel version")
            results_dataframe = results_dataframe[results_dataframe['kernelname'].apply(lambda x: LegacyVersion(x) > LegacyVersion(kernel_criteria))]
        elif kernel_criteria_op == "equals":
            logging.debug("Equals selected  for ", testname, " Kernel version")
            results_dataframe = results_dataframe[results_dataframe['kernelname'].apply(lambda x: LegacyVersion(x) == LegacyVersion(kernel_criteria))]
        else:
            logging.debug("Less than selected!  for ", testname, " Kernel version")
            results_dataframe = results_dataframe[results_dataframe['kernelname'].apply(lambda x: LegacyVersion(x) < LegacyVersion(kernel_criteria))]


    # Filter rows on os_version_criteria
    if os_version_criteria != '':
        if os_version_criteria_op == "greater than":
            logging.debug("greater than selected for ", testname, " OS version")
            results_dataframe = results_dataframe[results_dataframe['osversion'].apply(lambda x: LegacyVersion(x) > LegacyVersion(os_version_criteria))]
        elif kernel_criteria_op == "equals":
            logging.debug("Equals selected for ", testname, " OS version")
            results_dataframe = results_dataframe[results_dataframe['osversion'].apply(lambda x: LegacyVersion(x) == LegacyVersion(os_version_criteria))]
        else:
            logging.debug("Less than selected again!? for ", testname, " OS version")
            results_dataframe = results_dataframe[results_dataframe['osversion'].apply(lambda x: LegacyVersion(x) < LegacyVersion(os_version_criteria))]


    # Corresponding SKUID for skuidname
    # "Unkown SKUID" if skuidname not found in sku_definition.ini (For older benchmarking tests)
    if 'skuidname' in results_dataframe.columns:
        results_dataframe.insert(1, 'SKUID', [skuid_cpu_map.get(skuidname.strip(), "Unkown SKUID") for skuidname in results_dataframe['skuidname']])

    # Convert resultype column into corresponding entry of result_type_map
    if 'resultype' in results_dataframe.columns:
        index = list(results_dataframe.columns).index('resultype')
        results_dataframe.insert(index, 'Result Type', [result_type_map.get(result_type, "Unkown resultype") for result_type in results_dataframe['resultype']])

        # Drop the resultype column as it is no longer needed
        del results_dataframe['resultype']

    if best_results_condition:
        # Do NOT split the description column
        # Since it causes a LOT of columns to be displayed in the final Excel Sheet

        # Rename 'description' to 'Input file Description'
        results_dataframe = results_dataframe.rename(columns = {'description':'Input File Description'})
        pass
    else:
        # SPLIT the description string
        # Add Columns corresponding to the description_string
        description_list = description_string.split(',')
        description_list = list(map(lambda x: x.strip(), description_list))
        for col in reversed(description_list):
            results_dataframe.insert(len(results_dataframe.columns)-4, col, 'default value')

        # Function which splits the description string into various parameters 
        # according to 'description' field of the '.ini' file
        def split_description(index, description):
            try:
                return description.split(',')[index]
            except Exception as error_message:
                logging.debug("Error = {}".format(error_message))
                return np.nan

        # For all the rows in the dataframe, set the description_list values    
        for j in range(len(description_list)):
            results_dataframe[description_list[j]] = results_dataframe['description'].apply(lambda x: split_description(j, x))

        # Drop the 'description' column as we have now split it into various columns according to description_string
        del results_dataframe['description']

    # Add "FACTS Link" column at the end
    index = len(results_dataframe.columns)
    results_dataframe.insert(index, "FACTS Link", ['http://localhost:8005/test-details/' + str(originID) for originID in results_dataframe['originID']])

    return results_dataframe

@app.route('/generate_reports', methods=['POST'])
def generate_reports():
    db = pymysql.connect(host=DB_HOST_IP, user=DB_USER,
                         passwd=DB_PASSWD, db=DB_NAME, port=DB_PORT)

    logging.debug("Got request for generate reports")

    results_metadata_file_path = './config/best_of_all_graph.ini'
    results_metadata_parser = configparser.ConfigParser()
    results_metadata_parser.read(results_metadata_file_path)

    sku_file_path = './config/sku_definition.ini'
    sku_parser = configparser.ConfigParser()
    sku_parser.read(sku_file_path)

    #  Map from CPU Section->list(skuidname) in sku_definition.ini 
    skuid_cpu_map = OrderedDict({section: sku_parser.get(section, 'SKUID').replace('\"', '').split(',') for section in sku_parser.sections()})

    # Also a map from skuidname->Cpu Section. We need both in diff scenarios
    for section in sku_parser.sections():
        skuids = sku_parser.get(section, 'SKUID').replace('\"','').split(',')
        for skuid in skuids:
            skuid_cpu_map[skuid] = section

    # Map from filter label -> testname_list Eg. ncar -> ['clubb', 'mg2', 'waccm', 'wrf']
    label_testname_map = {}
    for section in results_metadata_parser.sections():
        labels = results_metadata_parser.get(section, 'label').replace('\"','').replace(' ', '').lower().split(',')

        # Remove empty strings
        labels = list(filter(None, labels))
        for label in labels:
            # If the list doesn't exist, create one
            if label not in label_testname_map:
                label_testname_map[label] = []
            label_testname_map[label].append(section)

    # Metadata 
    parameter_map = {
        "Kernel Version": 'os.kernelname', 
        'OS Version': 'os.osversion', 
        'OS Name': 'os.osdistro', 
        "Firmware Version": 'hw.fwversion' , 
        "ToolChain Name": 'tc.toolchainname', 
        "ToolChain Version" : 'tc.toolchainversion', 
        "Flags": 'tc.flags',
        "SMT" : 'b.smt',
        "Cores": 'b.cores',
        "Corefreq": 'b.corefreq',
        "DDRfreq": 'b.ddrfreq',
        "SKUID": 'n.skuidname',
        "Hostname": 'o.hostname',
        "Scaling" : 's.resultype',
        "Test Date" : 'o.testdate',
        "Notes" : 'o.notes',
    }

    # Required for date criteria
    month_name_number_map = {
        'Jan' : '01',
        'Feb' : '02',
        'Mar' : '03',
        'Apr' : '04',
        'May' : '05',
        'Jun' : '06',
        'Jul' : '07',
        'Aug' : '08',
        'Sep' : '09',
        'Oct' : '10',
        'Nov' : '11',
        'Dec' : '12',
    }


    param_list = [
        {'name' : 'Kernel Version', 'data_type' : 'string', 'display' : '', 'criteria' : '', 'criteria-op' : '', 'query_condition' : ''},
        {'name' : 'OS Version', 'data_type' : 'string', 'display' : '', 'criteria' : '', 'criteria-op' : '', 'query_condition' : ''},
        {'name' : 'OS Name', 'data_type' : 'string', 'display' : '', 'criteria' : '', 'criteria-op' : '', 'query_condition' : ''},
        {'name' : 'Firmware Version', 'data_type' : 'string', 'display' : '', 'criteria' : '','criteria-op' : '' , 'query_condition' : ''},
        {'name' : 'ToolChain Name', 'data_type' : 'string', 'display' : '', 'criteria' : '','criteria-op' : '' , 'query_condition' : ''},
        {'name' : 'ToolChain Version', 'data_type' : 'string', 'display' : '', 'criteria' : '','criteria-op' : '' , 'query_condition' : ''},
        {'name' : 'SMT', 'data_type' : 'numeric', 'display' : '', 'criteria' : '', 'criteria-op' : '', 'query_condition' : ''},
        {'name' : 'Cores', 'data_type' : 'numeric', 'display' : '', 'criteria' : '', 'criteria-op' : '', 'query_condition' : ''},
        {'name' : 'DDRfreq', 'data_type' : 'numeric', 'display' : '', 'criteria' : '', 'criteria-op' : '', 'query_condition' : ''},
        {'name' : 'SKUID', 'data_type' : 'string', 'display' : '', 'criteria' : '', 'criteria-op' : '', 'query_condition' : ''},
        {'name' : 'Hostname', 'data_type' : 'string', 'display' : '', 'criteria' : '', 'criteria-op' : '', 'query_condition' : ''},
        {'name' : 'Scaling', 'data_type' : 'string', 'display' : '', 'criteria' : '', 'criteria-op' : '', 'query_condition' : ''},
        {'name' : 'Test Date', 'data_type' : 'date', 'display' : '', 'criteria' : '', 'criteria2' : '', 'criteria-op' : '', 'query_condition' : ''},
        {'name' : 'Notes', 'data_type' : 'hidden', 'display' : '', 'criteria' : '', 'criteria-op' : '', 'query_condition' : ''},
    ]

    # Fill 'param_list' with the request params
    # Generate 'query_condition' for each param

    SELECT_PARAMS = " "
    FINAL_CRITERIA = " "

    # The criteria (key) which is used to check if excel file already exists in cached_results
    all_criteria_string = "reports-"

    # Get best_results_condition
    best_results_condition = request.form.get('best-results-radio')
    if not best_results_condition:
        # Empty string if value == None
        best_results_condition = ''
    logging.debug("GOT best results condition = \'{}\'".format(best_results_condition))

    all_criteria_string += best_results_condition + '-'

    # Append all criteria to the key
    all_criteria_string += "criteria-"

    skuid_criteria_op = ""
    all_skuidnames_criteria = []

    filename = request.form.get('filename')

    # Compute the metadata sheet parameters
    reports_metadata = {
        'Parameter' : ['Date Created'],
        'Option' : [''],
        'Value' : [filename[filename.find('-') + 2 : ]],
    }

    reports_metadata['Parameter'].append('Best Results')
    reports_metadata['Option'].append('')
    if best_results_condition:
        reports_metadata['Value'].append(best_results_condition)
    else:
        reports_metadata['Value'].append("No")

    for d in param_list:
        d['display'] = request.form.get('disp-'+d['name'])
        if d['name'] == 'SKUID':
            d['criteria'] = request.form.getlist('criteria-'+d['name'])
            # Remove the empty string '' from the list
            if '' in d['criteria']:
                d['criteria'].remove('')
        else:
            d['criteria'] = request.form.get('criteria-'+d['name'])
        d['criteria-op'] = request.form.get('criteria-op-'+d['name'])

        if d['criteria']:
            d['display'] = "Yes"

        if d['name'] == 'Test Date':
            d['criteria2'] = request.form.get('criteria2-'+d['name'])

        # Append to SELECT_PARAMS according to 'display' value
        if d['display'] == 'Yes':
            SELECT_PARAMS += parameter_map[d['name']] + ', '

        # Gives 'result type' number from result type string
        # Example if value="2 cores" then key=8
        def get_key_from_value(value):
            return str(list(result_type_map.keys())[list(result_type_map.values()).index(value)])

        # Append FINAL_CRITERIA
        if d['criteria']:
            reports_metadata['Parameter'].append(d['name'])
            reports_metadata['Option'].append(str(d['criteria-op']))
            reports_metadata['Value'].append(str(d['criteria']))

            if d['data_type'] == 'string':
                if d['name'] == 'Scaling':
                    if d['criteria-op'] == 'matches':
                        FINAL_CRITERIA += " AND s.resultype = " + get_key_from_value(d['criteria'].lower())
                    else:
                        FINAL_CRITERIA += " AND s.resultype = " + get_key_from_value(d['criteria'].lower())
                elif d['name'] == 'SKUID' and d['criteria'] != []:
                    for criteria in d['criteria']:
                        # Make list of lists for best_results_condition
                        # Since we have to get best results for each sku
                        if best_results_condition:
                            all_skuidnames_criteria.append(skuid_cpu_map[criteria])
                            skuid_criteria_op = d['criteria-op']
                        else:
                            all_skuidnames_criteria.extend(skuid_cpu_map[criteria])

                    # If not best_results_condition, append the criteria for skuidnames
                    if not best_results_condition:
                        if d['criteria-op'] == 'matches':
                            FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " IN " + str(all_skuidnames_criteria).replace('[','(').replace(']',')')
                        else:
                            FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " NOT IN " + str(all_skuidnames_criteria).replace('[','(').replace(']',')')

                elif d['name'] != 'Kernel Version' and d['name'] != 'OS Version':
                    if d['criteria-op'] == 'matches':
                        FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " LIKE \'%" + d['criteria'].strip() +"%\'"
                    else:
                        FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " NOT LIKE \'%" + d['criteria'].strip() +"%\'"
            elif d['data_type'] == 'numeric':
                if d['criteria-op'] == 'less than':
                    FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " < " + d['criteria'].strip()
                elif d['criteria-op'] == 'equals':
                    FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " = " + d['criteria'].strip()
                elif d['criteria-op'] == 'greater than':
                    FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " > " + d['criteria'].strip()
            elif d['data_type'] == 'date':
                if d['criteria-op'] == 'before':
                    FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " < \'" + d['criteria'].strip() + '-' + month_name_number_map[d['criteria2']] + '-' + '01' + "\'"
                elif d['criteria-op'] == 'during':
                    FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " LIKE \'%" + d['criteria'].strip() + '-' + month_name_number_map[d['criteria2']] + "-" "%\'"
                elif d['criteria-op'] == 'since':
                    FINAL_CRITERIA += " AND " + parameter_map[d['name']] + " > \'" + d['criteria'].strip() + '-' + month_name_number_map[d['criteria2']] + '-' + '01' + "\'"


    all_criteria_string += json.dumps(param_list)


    logging.debug(all_criteria_string)
    logging.debug(SELECT_PARAMS)
    logging.debug("ALL SKUIDNAMES = ", all_skuidnames_criteria)

    # Handle the condition where best_results_condition exists AND 'all_skuidnames_criteria' is empty []
    if best_results_condition and all_skuidnames_criteria == []:
        for criteria in sku_parser.sections():
            all_skuidnames_criteria.append(skuid_cpu_map[criteria])
        skuid_criteria_op = request.form.get('criteria-op-SKUID')

    # Retrieve data from the request object
    selected_sections_list = []
    if request.form.get('filter-by-label-or-benchmark') == 'testname':
        selected_sections_list = request.form.getlist('filter_testname_list')
    else:
        selected_labels_list = request.form.getlist('filter_labels_list')

        if selected_labels_list:
            reports_metadata['Parameter'].append("Selected Labels")
            reports_metadata['Option'].append('')
            reports_metadata['Value'].append(str(selected_labels_list))
            

        logging.debug("PRint selected_labels_list = {}".format(selected_labels_list))
        for label in selected_labels_list:
            selected_sections_list.extend(label_testname_map[label])

        # For getting unique entries
        selected_sections_list = list(set(selected_sections_list))

    # If none of the benchmarks is selected, put in all tests
    if selected_sections_list == []:
        selected_sections_list = results_metadata_parser.sections()

    # Sort the selected_tests_list alphabetically
    selected_sections_list.sort(key=str.lower)

    reports_metadata['Parameter'].append("Selected Benchmarks")
    reports_metadata['Option'].append('')
    reports_metadata['Value'].append(str(selected_sections_list))

    metadata_df = pd.DataFrame(reports_metadata)
    # Metadata dataframe is ready. Write it to the "Metadata" sheet while saving the file

    # Get the INPUT_FILTER_CONDITION for each selected test
    input_filters_list_list = [results_metadata_parser.get(test_section, 'default_input') \
                                .replace('\"', '').split(',') for test_section in selected_sections_list]
    INPUT_FILTER_CONDITION_LIST = [get_input_filter_condition(test_section, input_filters_list, \
                                    wiki_description_file="./config/best_of_all_graph.ini") \
                                    for test_section, input_filters_list in zip(selected_sections_list, input_filters_list_list)]

    # Read all test names each section of best_of_all_graph.ini file
    selected_tests_list = [results_metadata_parser.get(section, 'testname').strip() for section in selected_sections_list] 

    all_criteria_string += "selected-sections-" + json.dumps(selected_sections_list)
    # DONE! We got the all_criteria_string key 
    
    # The directory where cached excel files will be stored
    base_path = os.getcwd() + '/cached_results/'
    # If directory doesn't exist, create it
    if not os.path.exists(base_path):
        os.mkdir(base_path)         

    # cache_directory according to best-results variable
    if best_results_condition:
        cache_directory = base_path + 'reports_best_results/'
    else:
        cache_directory = base_path + 'reports_normal_results/'

    if not os.path.exists(cache_directory):
        os.mkdir(cache_directory)

    # Check if the excel file path for the current request exists in the map_file
    reports_cache_map_file = base_path + 'reports_cache_map.txt'

    # If reports_cache_map.txt doesn't exist, create it again
    if not os.path.isfile(reports_cache_map_file):
        os.system("touch " + reports_cache_map_file)
    
    # The actual dictionary. Fill it by using json.loads()
    reports_cache_map = {}
    with open(reports_cache_map_file, 'r') as f:
        try:
            reports_cache_map = json.loads(f.read())
        except:
            # In case the file is empty, ignore
            pass

    absolute_file_path = ""

    # Check if all_criteria_string exists in reports_cache_map's keys
    if all_criteria_string in reports_cache_map:
        logging.debug("Query entry found in Cache DICTIONARY")
        absolute_file_path = reports_cache_map[all_criteria_string]
    else:
        logging.debug("Query entry found in Cache DICTIONARY")
        reports_cache_map[all_criteria_string] = cache_directory + filename + '.xlsx'

    # If the file exists, then return it
    if os.path.exists(absolute_file_path):
        # Update the Metadata sheet
        wb = load_workbook(absolute_file_path)
        if 'Metadata' in wb.sheetnames:
            del wb['Metadata']

        # Create a new sheet. Insert at first position
        ws = wb.create_sheet("Metadata", 0)

        # Write the metadata dataframe to the "Metadata" sheet
        for r in dataframe_to_rows(metadata_df, index=True, header=True):
            ws.append(r)

        # Save the file, ready to return
        wb.save(absolute_file_path)
        logging.debug("Found cached file. Returning")
        pass
    else:
        logging.debug("File not found. Excecuting queries again")
        reports_cache_map[all_criteria_string] = cache_directory + filename + '.xlsx'

        # Filename .xlsx
        absolute_file_path = cache_directory + filename + '.xlsx'

        parallel_start_time = time.time()

        logging.debug("\n\nFilnalAll skuidnames criteria = {}".format(all_skuidnames_criteria))
        logging.debug("SKUID Criteria op = {}".format(skuid_criteria_op))

        # Parallel excecution 
        results_dataframe_list = []
        pool = multiprocessing.Pool(num_processes)
        try:
            results_dataframe_list = pool.map(partial(parallel_test_report, SELECT_PARAMS=SELECT_PARAMS, FINAL_CRITERIA=FINAL_CRITERIA, \
                                    kernel_criteria=request.form.get('criteria-Kernel Version'), os_version_criteria=request.form.get('criteria-OS Version'), \
                                    os_version_criteria_op=request.form.get('criteria-op-OS Version'), kernel_criteria_op=request.form.get('criteria-op-Kernel Version'), \
                                    skuid_cpu_map=skuid_cpu_map, best_results_condition=best_results_condition, skuid_criteria_op=skuid_criteria_op, \
                                    all_skuidnames_criteria=all_skuidnames_criteria), zip(selected_sections_list, selected_tests_list, INPUT_FILTER_CONDITION_LIST))
        except:
            pass
        finally:
            logging.debug("Closing Pool")
            pool.close()
            pool.join()

        logging.debug("Parallelism took {} seconds".format(time.time() - parallel_start_time))
        logging.debug("{}".format(len(results_dataframe_list)))
        start_time2 = time.time()

        if best_results_condition:
            logging.debug("Best results. Writing excel sheets")
            # Write all results in a single excel file

            final_results_dataframe = pd.DataFrame()

            # Append all the dataframes from the list to final_results_dataframe
            i = 0
            for results_dataframe in results_dataframe_list:
                i += 1
                final_results_dataframe = final_results_dataframe.append(results_dataframe, sort=False)

            # Reset Index
            final_results_dataframe = final_results_dataframe.reset_index(drop=True)

            logging.debug("final results dataframe = {}".format(i))

            # Write the metadata dataframe to create the excel file
            with pd.ExcelWriter(absolute_file_path, engine='openpyxl') as writer:
                metadata_df.to_excel(writer, sheet_name="Metadata")

            # Write the entire dataframe in a single sheet in append mode
            with pd.ExcelWriter(absolute_file_path, engine='openpyxl', mode='a') as writer:
                final_results_dataframe.to_excel(writer, sheet_name="Best results")

        else:
            logging.debug("Normal results. Writing excel sheets")
            # Write the metadata dataframe to create the excel file
            with pd.ExcelWriter(absolute_file_path, engine='openpyxl') as writer:
                metadata_df.to_excel(writer, sheet_name="Metadata")

            logging.debug("Wrote first excel sheet")
            
            # Write rest of the dataframes with the excel file in "Append" mode
            with pd.ExcelWriter(absolute_file_path, engine='openpyxl', mode='a') as writer:
                for results_dataframe, testname in zip(results_dataframe_list, selected_sections_list):
                    logging.debug("Writing excel sheet of {}".format(testname))
                    results_dataframe.to_excel(writer, sheet_name=testname)

        logging.debug("Writing all excel files took {} seconds".format(time.time() - start_time2))
    

    # Write the updated reports_cache_map_file back to the file
    with open(reports_cache_map_file, 'w') as f:
        f.write(json.dumps(reports_cache_map, indent=4))

    # Send the Excel file as response for download
    try:
        return send_file(absolute_file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            attachment_filename= filename + '.xlsx',
            as_attachment=True)

    except Exception as error_message:
        return error_message, 404

